# -*- coding: utf-8 -*-
"""
SoloDeportes vs Nike - Scraper Paralelo con Cache Inteligente
- Construye cache de SKU -> StyleColor desde PLPs
- Procesa PDPs desde cache en paralelo (4 workers)
- Maneja productos inactivos (404, sin stock, etc.)
- Reintenta 1 vez antes de marcar como inactivo
- Exporta CSV/XLSX con template exacto
"""

from __future__ import annotations

import os
import re
import json
import time
import math
import base64
import hashlib
import datetime as dt
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple, Set
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

import pandas as pd
from pyxlsb import open_workbook
from openai import OpenAI
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError


# =========================
# Config
# =========================
SEASON = os.getenv("SEASON", "SP26")

LINKS_FILE = os.getenv("LINKS_FILE", "Links Retail.xlsx")
LINKS_SHEET = os.getenv("LINKS_SHEET", "SoloDeportes")

STATUSBOOKS_FILE = os.getenv("STATUSBOOKS_FILE", "StatusBooks NDDC ARG SP26.xlsb")

# Límites
MAX_PRODUCTS_PER_PLP = 450
OMIT_FIRST_N = 4  # Omitir primeros 4 productos (más vendidos)
MAX_PARALLEL_WORKERS = 4  # Workers paralelos para PDPs
MAX_DAYS_ACTIVE = 30  # Procesar productos actualizados en últimos X días
MAX_RETRIES = 1  # Reintentos antes de marcar como inactivo

HEADLESS = os.getenv("HEADLESS", "false").lower() in ("1", "true", "yes", "y")

CACHE_FILE = os.getenv("SOLO_CACHE_FILE", "solodeportes_cache.json")

# OpenAI - CORREGIDO: Usar chat.completions no responses.create
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
OPENAI_TIMEOUT_S = int(os.getenv("OPENAI_TIMEOUT_S", "30"))
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "REMOVED_OPENAI_KEY")

# Shipping (placeholders)
SOLO_FREE_SHIP_FROM_ARS = 149999
SOLO_SHIPPING_ARS = 0
NIKE_FREE_SHIP_FROM_ARS = 229000
NIKE_SHIPPING_ARS = 10899


# =========================
# Helpers
# =========================
def now_ts() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")

def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def days_since(date_str: str) -> float:
    """Calcula días desde una fecha string"""
    if not date_str:
        return 999
    try:
        date_obj = dt.datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        return (dt.datetime.now() - date_obj).total_seconds() / 86400
    except:
        return 999

def money_to_int_ars(s: Optional[str]) -> Optional[int]:
    if not s:
        return None
    txt = str(s)
    txt = txt.replace("\xa0", " ").strip()
    m = re.search(r"\$?\s*([0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{2})?)", txt)
    if not m:
        return None
    num = m.group(1)
    num = num.replace(".", "")
    if "," in num:
        num = num.split(",")[0]
    try:
        return int(num)
    except Exception:
        return None

def pct_markdown(full: Optional[int], final: Optional[int]) -> Optional[float]:
    if full is None or final is None or full == 0:
        return None
    return (1.0 - (final / full)) * 100.0

def fmt_pct(x: Optional[float]) -> Optional[str]:
    if x is None:
        return None
    return f"{x:.1f}%"

def safe_float(x: Any) -> Optional[float]:
    try:
        if x is None or (isinstance(x, float) and math.isnan(x)):
            return None
        return float(x)
    except Exception:
        return None

def bml_pct(competitor: Optional[float], nike: Optional[float]) -> Optional[str]:
    if competitor is None or nike is None or nike == 0:
        return None
    pct = (competitor / nike - 1.0) * 100.0
    sign = "+" if pct > 0 else ""
    return f"{sign}{pct:.1f}%"

def load_json(path: str) -> Dict[str, Any]:
    default = {"by_sku": {}, "meta": {"created": now_str(), "version": 2}}
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        # Garantizar estructura correcta aunque el JSON esté vacío o corrupto
        if not isinstance(data, dict):
            return default
        if "by_sku" not in data:
            data["by_sku"] = {}
        if "meta" not in data:
            data["meta"] = {"created": now_str(), "version": 2}
        return data
    except Exception:
        return default

def save_json(path: str, data: Dict[str, Any]) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def normalize_url(url: str) -> str:
    """Normaliza URL a formato absoluto"""
    if not url:
        return ""
    
    url = url.strip()
    
    if url.startswith("http"):
        return url
    
    if url.startswith("/"):
        return f"https://www.solodeportes.com.ar{url}"
    
    return f"https://www.solodeportes.com.ar/{url}"

def is_valid_pdp_url(url: str) -> bool:
    """Verifica si es URL válida de PDP"""
    if not url:
        return False
    
    url_lower = url.lower()
    
    # DEBE terminar en .html
    if not url_lower.endswith(".html"):
        return False
    
    # NO DEBE tener parámetros de filtro/paginación
    exclude_patterns = [
        "?p=", "?genero=", "?type_product=", "?modelo=",
        "?order=", "?srule=", "?sz=", "?start=",
        "/marcas/", "marcas.html", "?cgid="
    ]
    
    for pattern in exclude_patterns:
        if pattern in url_lower:
            return False
    
    return True


# =========================
# StatusBooks - CORREGIDO
# =========================
@dataclass
class SBRow:
    stylecolor: str
    product_code: Optional[str] = None
    marketing_name: Optional[str] = None
    category: Optional[str] = None
    division: Optional[str] = None
    franchise: Optional[str] = None
    gender: Optional[str] = None
    nike_full_price: Optional[float] = None
    nike_final_price: Optional[float] = None
    nike_markdown: Optional[float] = None

def load_statusbooks_map(statusbooks_path: str) -> Dict[str, SBRow]:
    """Carga StatusBooks usando Product Code como StyleColor"""
    print("📚 Cargando Status Books...")
    
    STATUSBOOKS_SHEET = "Status Books NDDC SP26"
    HEADER_ROW_1BASED = 7  # Header en fila 7
    
    with open_workbook(statusbooks_path) as wb:
        # Encontrar el sheet correcto
        if STATUSBOOKS_SHEET not in wb.sheets:
            for sheet_name in wb.sheets:
                if "Status Books" in sheet_name or "NDDC" in sheet_name:
                    STATUSBOOKS_SHEET = sheet_name
                    break
            if STATUSBOOKS_SHEET not in wb.sheets:
                STATUSBOOKS_SHEET = wb.sheets[0]
                print(f"⚠️  Usando sheet: {STATUSBOOKS_SHEET}")
        
        print(f"   📄 Sheet: {STATUSBOOKS_SHEET}")
        
        with wb.get_sheet(STATUSBOOKS_SHEET) as sh:
            header_idx0 = HEADER_ROW_1BASED - 1
            header = None
            
            # Leer header (fila 7)
            for r_i, row in enumerate(sh.rows()):
                if r_i == header_idx0:
                    header = [str(c.v).strip() if c.v is not None else "" for c in row]
                    break
            
            if not header:
                raise ValueError(f"No pude leer header en fila {HEADER_ROW_1BASED}")
            
            # Debug: mostrar columnas encontradas
            print(f"   🔍 Columnas encontradas en fila {HEADER_ROW_1BASED}:")
            for idx, h in enumerate(header):
                if h:
                    print(f"     [{idx}] '{h}'")
            
            # Crear índice de columnas
            col_index = {}
            for idx, h in enumerate(header):
                if h:
                    col_index[h.strip().lower()] = idx
            
            def pick(*names: str) -> Optional[int]:
                # Primero intentar coincidencia exacta
                for n in names:
                    if n.lower() in col_index:
                        return col_index[n.lower()]
                
                # Búsqueda flexible (sin espacios, sin acentos)
                for n in names:
                    nl = n.lower().replace(" ", "").replace("_", "")
                    for k, v in col_index.items():
                        k_clean = k.replace(" ", "").replace("_", "")
                        if nl in k_clean or k_clean in nl:
                            return v
                return None
            
            # Buscar Product Code - nombres posibles
            ix_product_code = pick("Product Code", "ProductCode", "PRODUCT CODE", "Product_Code", 
                                 "Código Producto", "Código de Producto", "SKU", "Style", "Style-Color",
                                 "StyleColor", "Style Color", "Codigo Producto")
            
            if ix_product_code is None:
                # Último intento: buscar cualquier columna que contenga "product" o "code"
                for col_name, idx in col_index.items():
                    col_name_lower = col_name.lower()
                    if "product" in col_name_lower or "code" in col_name_lower or "sku" in col_name_lower or "style" in col_name_lower:
                        ix_product_code = idx
                        print(f"   ⚠️  Usando columna '{header[idx]}' como Product Code")
                        break
            
            if ix_product_code is None:
                print(f"   ❌ No se pudo encontrar columna 'Product Code'")
                print(f"   🔍 Columnas disponibles: {list(col_index.keys())}")
                raise ValueError("No encuentro columna de Product Code en StatusBooks.")
            
            print(f"   ✅ Columna Product Code: '{header[ix_product_code]}' (índice {ix_product_code})")
            
            # Buscar otras columnas
            ix_style = pick("Style", "STYLE", "Estilo", "Model")
            ix_mkt = pick("Marketing Name", "MarketingName", "Marketing name", "Nombre Marketing", 
                         "Description", "Descripción", "Product Name", "Nombre Producto")
            ix_cat = pick("Category", "Categoría", "Categorias", "Cat", "Categoria")
            ix_div = pick("Division", "BU", "Business Unit", "División", "Division")
            ix_fra = pick("Franchise", "Franquicia", "Línea", "Linea", "Family")
            ix_gen = pick("Gender", "Genero", "Género", "Sexo", "Target")
            ix_msrp = pick("VAL MSRP", "MSRP", "Precio Full", "Full Price", "Precio Original",
                          "Precio Lista", "List Price", "Price")
            ix_final = pick("VAL PRICE FINAL NUEVO", "VAL PRICE FINAL", "PRICE FINAL", 
                           "Precio Final", "Final Price", "Precio", "Selling Price")
            
            sb_map: Dict[str, SBRow] = {}
            rows_processed = 0
            
            # Procesar filas de datos
            for r_i, row in enumerate(sh.rows()):
                if r_i <= header_idx0:  # Saltar header y filas vacías
                    continue
                
                vals = [c.v for c in row]
                
                # Función helper para obtener valores
                def getv(ix):
                    if ix is None or ix >= len(vals):
                        return None
                    val = vals[ix]
                    if val is None:
                        return None
                    return str(val).strip() if isinstance(val, str) else val
                
                product_code = getv(ix_product_code)
                if product_code is None or product_code == "":
                    continue
                
                # Convertir a string y limpiar
                stylecolor = str(product_code).strip().upper()
                if not stylecolor or stylecolor.lower() in ["product code", "codigo", "code", "sku"]:
                    continue
                
                # Procesar precios
                nike_full = safe_float(getv(ix_msrp))
                nike_final = safe_float(getv(ix_final))
                nike_md = None
                if nike_full and nike_final is not None and nike_full != 0:
                    nike_md = 1 - (nike_final / nike_full)
                
                sb_map[stylecolor] = SBRow(
                    stylecolor=stylecolor,
                    product_code=str(getv(ix_style)).strip() if getv(ix_style) else "",
                    marketing_name=str(getv(ix_mkt)).strip() if getv(ix_mkt) else "",
                    category=str(getv(ix_cat)).strip() if getv(ix_cat) else "",
                    division=str(getv(ix_div)).strip() if getv(ix_div) else "",
                    franchise=str(getv(ix_fra)).strip() if getv(ix_fra) else "",
                    gender=str(getv(ix_gen)).strip() if getv(ix_gen) else "",
                    nike_full_price=nike_full,
                    nike_final_price=nike_final,
                    nike_markdown=nike_md
                )
                
                rows_processed += 1
                
                # Mostrar progreso cada 1000 filas
                if rows_processed % 1000 == 0:
                    print(f"   📊 {rows_processed:,} filas procesadas...")
                
                # Límite opcional para debug
                if rows_processed >= 10000:  # Límite para no sobrecargar
                    print(f"   ⚠️  Límite de 10,000 filas alcanzado")
                    break
    
    print(f"✅ StatusBooks OK: {len(sb_map):,} StyleColors procesados")
    return sb_map


# =========================
# Cache Management
# =========================
def update_cache_product(cache: Dict, sku: str, pdp_url: str, 
                        stylecolor: Optional[str] = None, 
                        is_active: bool = True,
                        error_reason: Optional[str] = None) -> None:
    """Actualiza o crea producto en cache"""
    if "by_sku" not in cache:
        cache["by_sku"] = {}
    
    if sku not in cache["by_sku"]:
        cache["by_sku"][sku] = {
            "sku": sku,
            "pdp_url": pdp_url,
            "stylecolor": stylecolor,
            "active": is_active,
            "created": now_str(),
            "last_updated": now_str(),
            "last_checked": now_str(),
            "retry_count": 0,
            "error_reason": error_reason,
            "product_name": "",
            "brand": "nike"
        }
    else:
        cache["by_sku"][sku]["last_checked"] = now_str()
        cache["by_sku"][sku]["retry_count"] = cache["by_sku"][sku].get("retry_count", 0)
        
        if error_reason:
            cache["by_sku"][sku]["error_reason"] = error_reason
            cache["by_sku"][sku]["retry_count"] += 1
            
            # Si supera retries, marcar como inactivo
            if cache["by_sku"][sku]["retry_count"] > MAX_RETRIES:
                cache["by_sku"][sku]["active"] = False
        else:
            cache["by_sku"][sku]["error_reason"] = None
            cache["by_sku"][sku]["retry_count"] = 0
            cache["by_sku"][sku]["active"] = is_active
        
        if stylecolor:
            cache["by_sku"][sku]["stylecolor"] = stylecolor
            cache["by_sku"][sku]["last_updated"] = now_str()
        
        if pdp_url:
            cache["by_sku"][sku]["pdp_url"] = pdp_url
    
    # Actualizar metadata
    cache["meta"] = {
        "last_updated": now_str(),
        "total_products": len(cache["by_sku"]),
        "active_products": sum(1 for p in cache["by_sku"].values() if p.get("active", True)),
        "version": 2
    }

def get_active_recent_products(cache: Dict, max_days: int = MAX_DAYS_ACTIVE) -> Dict[str, Dict]:
    """Obtiene productos activos y recientes del cache"""
    active_products = {}
    
    for sku, product in cache.get("by_sku", {}).items():
        # Verificar activo
        if not product.get("active", True):
            continue
        
        # Verificar que tenga stylecolor
        if not product.get("stylecolor"):
            continue
        
        # Verificar que sea reciente
        last_updated = product.get("last_updated", product.get("created", ""))
        if days_since(last_updated) > max_days:
            continue
        
        active_products[sku] = product
    
    return active_products

def find_sku_by_url(cache: Dict, url: str) -> Optional[str]:
    """Encuentra SKU por URL en cache"""
    url_normalized = normalize_url(url)
    
    for sku, product in cache.get("by_sku", {}).items():
        product_url = normalize_url(product.get("pdp_url", ""))
        if product_url == url_normalized:
            return sku
    
    return None


# =========================
# OpenAI Functions - ¡MODIFICADO! (SKU como texto)
# =========================
def ask_stylecolor_from_sku_safe(client: Optional[OpenAI], sku_solodeportes: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Extrae el StyleColor Nike directamente del SKU de SoloDeportes sin usar OpenAI.
    El SKU tiene estructura: PREFIJO(6 dígitos) + STYLE(2 letras + 4 dígitos) + COLOR(3 dígitos)
    Ejemplo: 510010FQ8146002 → FQ8146-002
    Devuelve: (stylecolor_nike, error_message)
    """
    if not sku_solodeportes or not sku_solodeportes.strip():
        return None, "SKU vacío"

    sku = sku_solodeportes.strip().upper()

    # Patrón principal: 6 dígitos + 2 letras + 4 dígitos + 3 dígitos
    m = re.match(r"^\d{6}([A-Z]{2}\d{4})(\d{3})$", sku)
    if m:
        stylecolor = f"{m.group(1)}-{m.group(2)}"
        print(f"      ✅ StyleColor extraído del SKU: {stylecolor}")
        return stylecolor, None

    # Patrón alternativo: 6 dígitos + letras+dígitos mezclados + 3 dígitos al final
    m = re.match(r"^\d{6}([A-Z]{2}\d{3,5}[A-Z0-9]{0,3})(\d{3})$", sku)
    if m:
        stylecolor = f"{m.group(1)}-{m.group(2)}"
        print(f"      ✅ StyleColor extraído (alt): {stylecolor}")
        return stylecolor, None

    # Patrón para SKUs con letras intermedias: ej 39501332C2UW0KB
    m = re.match(r"^\d{5,6}([A-Z0-9]{6,8})([A-Z0-9]{3})$", sku)
    if m:
        candidate = f"{m.group(1)}-{m.group(2)}"
        # Validar que tenga al menos 2 letras al inicio
        if re.match(r"^[A-Z]{2}", m.group(1)):
            print(f"      ✅ StyleColor extraído (flex): {candidate}")
            return candidate, None

    print(f"      ⚠️  No se pudo extraer StyleColor del SKU: {sku}")
    return None, f"SKU sin patrón Nike reconocible: {sku}"

def _ask_stylecolor_from_sku_safe_UNUSED(client, sku_solodeportes):
    """OBSOLETO — reemplazado por extracción directa de SKU"""
    # Mantenido solo como referencia, no se usa
    try:
        pass
    except Exception as e:
        error_msg = str(e)
        print(f"      🔍 Error OpenAI: {error_msg}")
        if "401" in error_msg or "authentication" in error_msg.lower():
            return None, "API key inválida o expirada"
        elif "429" in error_msg:
            return None, "Rate limit excedido"
        elif "quota" in error_msg.lower():
            return None, "Quota excedido"
        elif "timeout" in error_msg.lower():
            return None, "Timeout en consulta"
        else:
            return None, f"Error: {error_msg[:100]}"

def normalize_stylecolor_for_match(style: str) -> List[str]:
    """Genera variantes para matching con StatusBooks"""
    style = (style or "").strip().upper()
    style = style.replace(" ", "")
    variants = []
    
    if style:
        variants.append(style)
        variants.append(style.replace("-", ""))
        if "-" not in style and len(style) >= 9:
            variants.append(style[:-3] + "-" + style[-3:])
    
    # Unique preserving order
    out = []
    for v in variants:
        if v and v not in out:
            out.append(v)
    
    return out


# =========================
# PLP Scraping (Cache Building) - Mejorado
# =========================
def read_links_excel(path: str, sheet: str) -> List[str]:
    """Lee URLs de PLPs desde Excel"""
    if not os.path.exists(path):
        raise FileNotFoundError(f"No encuentro {path}")
    
    df = pd.read_excel(path, sheet_name=sheet)
    
    # Buscar columna con URLs
    url_col = None
    for c in df.columns:
        col_lower = str(c).lower()
        if "http" in col_lower or "url" in col_lower or "link" in col_lower:
            url_col = c
            break
    
    if url_col is None:
        for c in df.columns:
            if df[c].astype(str).str.contains("http", na=False).any():
                url_col = c
                break
    
    if url_col is None:
        raise ValueError(f"No encuentro columna con URLs en {path} sheet {sheet}.")
    
    urls = df[url_col].dropna().astype(str).tolist()
    return [u.strip() for u in urls if u.strip().startswith("http")]

def click_load_more_plp(page, max_clicks: int = 10) -> int:
    """Intenta cargar más productos en PLP"""
    clicks_done = 0
    
    load_more_patterns = [
        "button:has-text('Ver más')",
        "button:has-text('Cargar más')",
        "button:has-text('Mostrar más')",
        "a:has-text('Ver más')",
        "a:has-text('Cargar más')",
        "[class*='load-more']",
        "[class*='view-more']",
    ]
    
    for _ in range(max_clicks):
        clicked = False
        
        for pattern in load_more_patterns:
            try:
                btn = page.locator(pattern).first
                if btn.count() > 0 and btn.is_visible():
                    btn.scroll_into_view_if_needed()
                    btn.click(timeout=3000)
                    page.wait_for_timeout(2500)
                    clicks_done += 1
                    clicked = True
                    print(f"      🔄 Click {clicks_done} en botón")
                    break
            except Exception:
                continue
        
        if not clicked:
            break
    
    return clicks_done

def simulate_scroll_load(page, scrolls: int = 10):
    """Simula scroll para cargar productos automáticamente"""
    for i in range(scrolls):
        try:
            page.evaluate("window.scrollBy(0, 1000)")
            page.wait_for_timeout(2000)
            
            # Verificar si hay nuevos productos
            product_count = page.locator("[class*='product'], .product-item").count()
            print(f"      Scroll {i+1}: {product_count} productos")
            
        except Exception:
            break

def extract_sku_from_product_container(container) -> Optional[str]:
    """Extrae SKU del contenedor de producto en PLP"""
    # Método 1: Buscar .product-item-sku .value
    try:
        sku_elem = container.locator(".product-item-sku .value").first
        if sku_elem.count() > 0:
            sku = sku_elem.inner_text(timeout=500).strip()
            if sku:
                return sku
    except Exception:
        pass
    
    # Método 2: Buscar en data attributes
    try:
        html = container.inner_html()
        match = re.search(r'<span class="value">([A-Z0-9-]+)</span>', html)
        if match:
            return match.group(1).strip()
    except Exception:
        pass
    
    return None

def extract_pdp_url_from_container(container) -> Optional[str]:
    """Extrae URL de PDP del contenedor de producto"""
    # Buscar enlaces .product-item-link o .product-item-photo
    selectors = [
        "a.product-item-link",
        ".product-item-photo a",
        ".product-item-info a",
        "a[href$='.html']",
    ]
    
    for selector in selectors:
        try:
            link = container.locator(selector).first
            if link.count() > 0:
                href = link.get_attribute("href")
                if href and is_valid_pdp_url(href):
                    return normalize_url(href)
        except Exception:
            continue
    
    return None

def extract_products_from_plp(page, cache: Dict, client: Optional[OpenAI] = None) -> int:
    """
    Extrae productos de PLP y actualiza cache.
    Devuelve número de productos nuevos encontrados.
    """
    print("   🔍 Extrayendo productos de PLP...")
    
    # Cargar más productos
    simulate_scroll_load(page, scrolls=8)
    click_load_more_plp(page)
    
    # Encontrar contenedores de producto
    product_containers = page.locator(".product-item, [class*='product'], .item")
    total_count = product_containers.count()
    
    print(f"   📊 Contenedores encontrados: {total_count}")
    
    new_products = 0
    browser_connected = True
    
    # Procesar cada producto (omitir primeros N)
    for i in range(OMIT_FIRST_N, min(total_count, MAX_PRODUCTS_PER_PLP + OMIT_FIRST_N)):
        if not browser_connected:
            print("      ⚠️  Browser desconectado, deteniendo extracción")
            break
            
        try:
            container = product_containers.nth(i)
            
            # Extraer SKU
            sku = extract_sku_from_product_container(container)
            if not sku:
                continue
            
            # Extraer URL PDP
            pdp_url = extract_pdp_url_from_container(container)
            if not pdp_url:
                continue
            
            # Verificar si ya está en cache
            if sku in cache["by_sku"]:
                # Si está inactivo pero lo encontramos en PLP, reactivar
                if not cache["by_sku"][sku].get("active", True):
                    print(f"      🔄 Reactivando SKU {sku} (encontrado en PLP)")
                    cache["by_sku"][sku]["active"] = True
                    cache["by_sku"][sku]["retry_count"] = 0
                    cache["by_sku"][sku]["error_reason"] = None
                
                # Ya procesado, skip
                continue
            
            # NUEVO PRODUCTO: Necesitamos StyleColor via OpenAI
            print(f"      🆕 Nuevo producto: SKU {sku}")
            
            # Navegar a PDP para screenshot
            try:
                page.goto(pdp_url, wait_until="domcontentloaded", timeout=30000)
                page.wait_for_timeout(2000)
                
                # Tomar screenshot MÁS PEQUEÑO (área específica)
                # Buscar área principal del producto
                try:
                    product_area = page.locator(".product-media, .product-image, .product-info-main, .product.media")
                    if product_area.count() > 0:
                        png = product_area.first.screenshot()
                    else:
                        # Fallback: screenshot de ventana (recortado)
                        png = page.screenshot(full_page=False, clip={"x": 0, "y": 0, "width": 1200, "height": 800})
                except Exception as screenshot_error:
                    print(f"         ⚠️  Error screenshot: {screenshot_error}")
                    png = page.screenshot(full_page=False)
                
                # Consultar OpenAI (con manejo seguro) - ¡MODIFICADO!
                stylecolor = None
                error_msg = None
                
                if client:
                    # Usamos el SKU de SoloDeportes (ya extraído de la PLP) como texto
                    stylecolor, error_msg = ask_stylecolor_from_sku_safe(client, sku)
                    # Pequeña pausa para evitar rate limiting
                    time.sleep(1.5)
                
                # Volver a PLP (SOLO si el browser sigue conectado)
                try:
                    if page.context and page.context.browser and page.context.browser.is_connected():
                        page.go_back()
                        page.wait_for_timeout(1500)
                    else:
                        browser_connected = False
                        print("      ⚠️  Browser desconectado durante procesamiento")
                except Exception as e:
                    print(f"      ⚠️  Error volviendo a PLP: {e}")
                    browser_connected = False
                
                # Guardar en cache
                update_cache_product(
                    cache=cache,
                    sku=sku,
                    pdp_url=pdp_url,
                    stylecolor=stylecolor,
                    is_active=stylecolor is not None,  # Activo solo si tiene StyleColor
                    error_reason=error_msg
                )
                
                new_products += 1
                
                if stylecolor:
                    print(f"         ✅ StyleColor: {stylecolor}")
                else:
                    print(f"         ⚠️  Sin StyleColor: {error_msg or 'OpenAI no devolvió resultado'}")
                
            except Exception as e:
                error_str = str(e)
                print(f"         ❌ Error procesando PDP: {error_str[:100]}")
                
                # Verificar si es error de navegación
                if "closed" in error_str.lower() or "disconnected" in error_str.lower():
                    browser_connected = False
                    print("      ⚠️  Browser cerrado/desconectado")
                
                # Guardar igual pero marcado como error
                update_cache_product(
                    cache=cache,
                    sku=sku,
                    pdp_url=pdp_url,
                    stylecolor=None,
                    is_active=False,
                    error_reason=error_str[:200]
                )
            
        except Exception as e:
            print(f"      ⚠️  Error procesando producto {i}: {e}")
            continue
    
    return new_products

def build_cache_from_plps(cache: Dict) -> int:
    """
    Construye/actualiza cache navegando PLPs.
    Devuelve total de productos nuevos.
    """
    print("\n🔨 Construyendo/Actualizando Cache desde PLPs...")
    
    # Leer PLPs
    plp_urls = read_links_excel(LINKS_FILE, LINKS_SHEET)
    print(f"📊 PLPs a procesar: {len(plp_urls)}")
    
    # Inicializar OpenAI (pero continuar incluso si falla)
    client = None
    openai_error = None
    
    try:
        if OPENAI_API_KEY and "sk-" in OPENAI_API_KEY:
            client = OpenAI(api_key=OPENAI_API_KEY)
            print("   🤖 OpenAI inicializado (usando chat.completions)")
            
            # Prueba rápida de conexión
            try:
                test_response = client.chat.completions.create(
                    model=OPENAI_MODEL,
                    messages=[{"role": "user", "content": "Test"}],
                    max_tokens=5
                )
                print("   ✅ Conexión OpenAI verificada")
            except Exception as test_error:
                print(f"   ⚠️  Advertencia en test OpenAI: {test_error}")
        else:
            print("   ⚠️  API key de OpenAI no válida o faltante")
    except Exception as e:
        openai_error = str(e)
        print(f"   ❌ Error inicializando OpenAI: {openai_error}")
        print("   ⚠️  Continuando sin OpenAI (solo extraeremos URLs)")
    
    total_new = 0
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        page = context.new_page()
        
        for plp_idx, plp_url in enumerate(plp_urls, 1):
            print(f"\n📦 PLP {plp_idx}/{len(plp_urls)}: {plp_url}")
            
            try:
                page.goto(plp_url, wait_until="domcontentloaded", timeout=45000)
                page.wait_for_timeout(3000)
                
                # Extraer productos
                new_in_plp = extract_products_from_plp(page, cache, client)
                total_new += new_in_plp
                
                print(f"   ✅ Nuevos en este PLP: {new_in_plp}")
                
                # Guardar cache periódicamente
                if plp_idx % 3 == 0:
                    save_json(CACHE_FILE, cache)
                    print(f"   💾 Cache guardado ({len(cache['by_sku'])} productos)")
                
                # Pausa entre PLPs
                time.sleep(2)
                
            except Exception as e:
                print(f"   ❌ Error en PLP: {e}")
                continue
        
        browser.close()
    
    # Guardar cache final
    if "by_sku" not in cache:
        cache["by_sku"] = {}
    save_json(CACHE_FILE, cache)
    
    print(f"\n✅ Cache construido: {len(cache['by_sku'])} productos totales")
    print(f"📊 Productos nuevos agregados: {total_new}")
    
    if openai_error:
        print(f"⚠️  Advertencia: OpenAI tuvo errores: {openai_error}")
        print("💡 Algunos productos pueden no tener StyleColor")
    
    return total_new


# =========================
# PDP Processing (Parallel)
# =========================
def check_pdp_active_and_extract(page, pdp_url: str) -> Tuple[bool, Optional[Dict], Optional[str]]:
    """
    Verifica si PDP está activa y extrae datos.
    Devuelve (is_active, data_dict, error_reason)
    """
    try:
        # Navegar a PDP
        response = page.goto(pdp_url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(1500)
        
        # Verificar respuesta HTTP
        if response and response.status >= 400:
            return False, None, f"HTTP {response.status}"
        
        # Verificar texto de error
        page_text = page.inner_text("body").lower()
        error_indicators = [
            "producto no disponible",
            "no encontrado",
            "error 404",
            "out of stock",
            "sin stock",
            "discontinuado",
            "agotado",
            "página no encontrada",
        ]
        
        for indicator in error_indicators:
            if indicator in page_text:
                return False, None, f"Texto: {indicator}"
        
        # Extraer precios
        full_price, final_price, markdown = extract_prices_from_pdp(page)
        
        # Verificar que haya al menos un precio
        if final_price is None:
            return False, None, "Sin precios visibles"
        
        # Extraer cuotas
        cuotas, tiene_interes = extract_cuotas_from_pdp(page)
        
        # Extraer SKU (verificación adicional)
        sku = extract_sku_from_pdp(page)
        
        return True, {
            "full_price": full_price,
            "final_price": final_price,
            "markdown": markdown,
            "cuotas": cuotas,
            "tiene_interes": tiene_interes,
            "sku": sku,
            "url": pdp_url
        }, None
        
    except Exception as e:
        return False, None, f"Excepción: {str(e)}"

def extract_prices_from_pdp(page) -> Tuple[Optional[int], Optional[int], Optional[str]]:
    """Extrae precios desde PDP"""
    final_price = None
    full_price = None
    
    # Buscar precio final
    final_selectors = [
        "[data-price-type='finalPrice'] .price",
        ".price-final_price .price",
        ".special-price .price",
        ".product-info-price .price-box .price",
        ".sales .price",
        "span.price:not(.old-price) span.price",
    ]
    
    for sel in final_selectors:
        try:
            element = page.locator(sel).first
            if element.count() > 0:
                text = element.inner_text(timeout=1500)
                price = money_to_int_ars(text)
                if price:
                    final_price = price
                    break
        except Exception:
            continue
    
    # Buscar precio original
    full_selectors = [
        "[data-price-type='oldPrice'] .price",
        ".old-price .price",
        ".price-old",
        "del .price",
        ".price-through .price",
    ]
    
    for sel in full_selectors:
        try:
            element = page.locator(sel).first
            if element.count() > 0:
                text = element.inner_text(timeout=1500)
                price = money_to_int_ars(text)
                if price:
                    full_price = price
                    break
        except Exception:
            continue
    
    # Si no hay full, usar final
    if full_price is None and final_price is not None:
        full_price = final_price
    
    # Calcular markdown
    markdown = None
    if full_price and final_price and full_price > 0 and full_price != final_price:
        discount_pct = ((full_price - final_price) / full_price) * 100
        markdown = f"{discount_pct:.1f}%"
    
    return full_price, final_price, markdown

def extract_cuotas_from_pdp(page) -> Tuple[Optional[int], Optional[bool]]:
    """Extrae cuotas desde PDP"""
    try:
        html = page.content()
        
        # Buscar sin interés - CORREGIDO: usar raw string
        sin_interes_patterns = [
            r'(\d+)\s*cuotas\s*sin\s*inter[eé]s',
            r'sin\s*inter[eé]s\s*en\s*(\d+)\s*cuotas',
            r'(\d+)\s*cuotas\s*s/i',
            r'(\d+)\s*x\s*sin\s*inter[eé]s',
        ]
        
        for pattern in sin_interes_patterns:
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                try:
                    cuotas = int(match.group(1))
                    return cuotas, False  # sin interés
                except:
                    continue
        
        # Buscar cualquier cuota
        match = re.search(r'(\d+)\s*cuotas', html, re.IGNORECASE)
        if match:
            try:
                cuotas = int(match.group(1))
                return cuotas, True  # con interés (asumido)
            except:
                pass
                
    except Exception:
        pass
    
    return None, None

def extract_sku_from_pdp(page) -> Optional[str]:
    """Extrae SKU desde PDP (verificación)"""
    selectors = [
        "div.product.attribute.sku div.value",
        "[itemprop='sku']",
        ".sku .value",
        ".product-info-stock-sku .sku",
    ]
    
    for sel in selectors:
        try:
            element = page.locator(sel).first
            if element.count() > 0:
                text = element.inner_text(timeout=1000).strip()
                if text:
                    return re.sub(r'\s+', '', text)
        except Exception:
            continue
    
    return None

def worker_process_pdps(urls_batch: List[str], worker_id: int) -> List[Dict]:
    """
    Worker que procesa un batch de PDPs.
    Cada worker tiene su propio contexto de Playwright.
    """
    worker_results = []
    
    print(f"   👷 Worker {worker_id} iniciando ({len(urls_batch)} URLs)")
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        
        for url_idx, url in enumerate(urls_batch):
            try:
                page = context.new_page()
                
                is_active, data, error = check_pdp_active_and_extract(page, url)
                
                result = {
                    "url": url,
                    "success": is_active,
                    "data": data if is_active else None,
                    "error": error if not is_active else None,
                    "worker_id": worker_id,
                    "url_index": url_idx
                }
                
                worker_results.append(result)
                
                page.close()
                
                # Pequeña pausa entre requests
                time.sleep(0.5)
                
            except Exception as e:
                worker_results.append({
                    "url": url,
                    "success": False,
                    "data": None,
                    "error": f"Worker exception: {str(e)}",
                    "worker_id": worker_id,
                    "url_index": url_idx
                })
        
        browser.close()
    
    print(f"   ✅ Worker {worker_id} completado")
    return worker_results

def process_pdps_parallel(pdp_urls: List[str], max_workers: int = MAX_PARALLEL_WORKERS) -> List[Dict]:
    """
    Procesa múltiples PDPs en paralelo.
    """
    print(f"\n🚀 Procesando {len(pdp_urls)} PDPs en paralelo ({max_workers} workers)...")
    
    # Dividir URLs en batches
    batch_size = max(1, len(pdp_urls) // max_workers)
    batches = [pdp_urls[i:i + batch_size] for i in range(0, len(pdp_urls), batch_size)]
    
    # Limitar batches a max_workers
    batches = batches[:max_workers]
    
    all_results = []
    
    with ThreadPoolExecutor(max_workers=len(batches)) as executor:
        # Enviar cada batch a un worker
        future_to_batch = {
            executor.submit(worker_process_pdps, batch, i): (batch, i)
            for i, batch in enumerate(batches)
        }
        
        # Recolectar resultados
        for future in as_completed(future_to_batch):
            batch, worker_id = future_to_batch[future]
            
            try:
                results = future.result(timeout=300)
                all_results.extend(results)
                
                success_count = sum(1 for r in results if r["success"])
                print(f"   📊 Worker {worker_id}: {success_count}/{len(results)} exitosas")
                
            except Exception as e:
                print(f"   ❌ Worker {worker_id} error: {e}")
                
                # Marcar todas las URLs de este batch como fallidas
                for url in batch:
                    all_results.append({
                        "url": url,
                        "success": False,
                        "data": None,
                        "error": f"Worker timeout/error: {str(e)}",
                        "worker_id": worker_id,
                        "url_index": -1
                    })
    
    # Estadísticas
    success_count = sum(1 for r in all_results if r["success"])
    print(f"\n📈 Resultados: {success_count}/{len(all_results)} PDPs exitosas")
    
    return all_results


# =========================
# Output Generation
# =========================
def build_template_df() -> pd.DataFrame:
    """Crea DataFrame con template de salida"""
    cols = [
        "StyleColor", "ProductCode", "MarketingName", "Category", "Division",
        "Franchise", "Gender", "Retailer", "Retailer_PDP", "Retailer_FullPrice",
        "Retailer_FinalPrice", "Retailer_Markdown", "Nike_FullPrice", "Nike_FinalPrice",
        "Nike_Markdown", "BML_FinalPrice", "BML_FullPrice", "Retailer_Shipping",
        "Nike_Shipping", "Nike_Price+Shipping", "Retailer_Price+Shipping",
        "BML_with_Shipping", "Cuotas_Retailer", "Cuotas_Nike", "BML_Cuotas",
        "Season", "Run_Date", "Last_Update_Retailer", "Is_Football", "Price_Source",
        "Raw_Nike_PriceText", "Raw_Retailer_PriceText"
    ]
    return pd.DataFrame(columns=cols)

def create_output_row(cache_product: Dict, pdp_data: Dict, sb_row: SBRow) -> Dict:
    """Crea una fila para el output"""
    # Shipping calculations
    comp_ship = 0 if (pdp_data["final_price"] or 0) >= SOLO_FREE_SHIP_FROM_ARS else SOLO_SHIPPING_ARS
    nike_ship = 0 if (sb_row.nike_final_price or 0) >= NIKE_FREE_SHIP_FROM_ARS else NIKE_SHIPPING_ARS
    
    nike_price_ship = (sb_row.nike_final_price + nike_ship) if sb_row.nike_final_price is not None else None
    comp_price_ship = (pdp_data["final_price"] + comp_ship) if pdp_data["final_price"] is not None else None
    
    # BML calculations
    bml_final = bml_pct(pdp_data["final_price"], sb_row.nike_final_price)
    bml_full = bml_pct(pdp_data["full_price"], sb_row.nike_full_price)
    bml_ship = bml_pct(comp_price_ship, nike_price_ship) if comp_price_ship and nike_price_ship else None
    
    # Cuotas (solo retailer por ahora)
    cuotas_nike = None
    bml_cuotas = None
    
    return {
        "StyleColor": sb_row.stylecolor,
        "ProductCode": sb_row.product_code,
        "MarketingName": sb_row.marketing_name,
        "Category": sb_row.category,
        "Division": sb_row.division,
        "Franchise": sb_row.franchise,
        "Gender": sb_row.gender,
        "Retailer": "SoloDeportes",
        "Retailer_PDP": cache_product["pdp_url"],
        "Retailer_FullPrice": pdp_data["full_price"],
        "Retailer_FinalPrice": pdp_data["final_price"],
        "Retailer_Markdown": pdp_data["markdown"],
        "Nike_FullPrice": sb_row.nike_full_price,
        "Nike_FinalPrice": sb_row.nike_final_price,
        "Nike_Markdown": fmt_pct(sb_row.nike_markdown * 100) if sb_row.nike_markdown else None,
        "BML_FinalPrice": bml_final,
        "BML_FullPrice": bml_full,
        "Retailer_Shipping": comp_ship,
        "Nike_Shipping": nike_ship,
        "Nike_Price+Shipping": nike_price_ship,
        "Retailer_Price+Shipping": comp_price_ship,
        "BML_with_Shipping": bml_ship,
        "Cuotas_Retailer": pdp_data.get("cuotas"),
        "Cuotas_Nike": cuotas_nike,
        "BML_Cuotas": bml_cuotas,
        "Season": SEASON,
        "Run_Date": now_str(),
        "Last_Update_Retailer": now_str(),
        "Is_Football": None,
        "Price_Source": "SoloDeportes PDP + StatusBooks",
        "Raw_Nike_PriceText": None,
        "Raw_Retailer_PriceText": None,
    }

def export_xlsx(df: pd.DataFrame, path: str) -> None:
    """Exporta DataFrame a XLSX con formato"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    ws = wb.active
    ws.title = "SoloDeportes_vs_Nike"
    
    header_fill = PatternFill("solid", fgColor="F15A24")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=1, column=c_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    
    for col in ws.columns:
        max_len = 10
        col_letter = col[0].column_letter
        for cell in col[:200]:
            if cell.value is not None:
                max_len = max(max_len, min(60, len(str(cell.value))))
        ws.column_dimensions[col_letter].width = max_len + 2
    
    wb.save(path)


# =========================
# Main Execution
# =========================
def main():
    print("=" * 70)
    print("🚀 SOLODEPORTES vs NIKE - Scraper Paralelo con Cache")
    print(f"📅 Season: {SEASON}")
    print(f"👷 Workers paralelos: {MAX_PARALLEL_WORKERS}")
    print(f"📊 Días activos: {MAX_DAYS_ACTIVE}")
    print("=" * 70)
    
    # 1. Cargar StatusBooks
    print("\n1️⃣ Cargando StatusBooks...")
    sb_map = load_statusbooks_map(STATUSBOOKS_FILE)
    
    # 2. Cargar Cache
    print("\n2️⃣ Cargando Cache...")
    cache = load_json(CACHE_FILE)
    print(f"   📊 Cache actual: {len(cache.get('by_sku', {}))} productos")
    print(f"   ✅ Activos: {sum(1 for p in cache.get('by_sku', {}).values() if p.get('active', True))}")
    
    # 3. Construir/Actualizar Cache si es necesario
    active_count = sum(1 for p in cache.get("by_sku", {}).values() if p.get("active", True))
    
    if active_count < 100:  # Umbral bajo, necesitamos construir cache
        print(f"\n3️⃣ ⚠️  Cache pequeño ({active_count} activos), construyendo...")
        build_cache_from_plps(cache)
    else:
        print(f"\n3️⃣ ✅ Cache suficiente ({active_count} activos), saltando construcción")
    
    # 4. Obtener productos activos y recientes para procesar
    print(f"\n4️⃣ Filtrando productos activos (últimos {MAX_DAYS_ACTIVE} días)...")
    active_products = get_active_recent_products(cache, max_days=MAX_DAYS_ACTIVE)
    
    if not active_products:
        print("   ⚠️  No hay productos activos recientes para procesar")
        print("   💡 Ejecuta nuevamente para procesar productos del cache")
        
        # Mostrar estadísticas del cache
        products_with_style = sum(1 for p in cache.get("by_sku", {}).values() if p.get("stylecolor"))
        products_without_style = sum(1 for p in cache.get("by_sku", {}).values() if not p.get("stylecolor"))
        
        print(f"   📊 Diagnóstico cache:")
        print(f"      • Con StyleColor: {products_with_style}")
        print(f"      • Sin StyleColor: {products_without_style}")
        
        if products_without_style > 0:
            print(f"\n   💡 Recomendación: OpenAI funciona, ejecuta nuevamente para obtener StyleColor")
        
        return
    
    print(f"   📊 Productos a procesar: {len(active_products)}")
    
    # 5. Extraer URLs de PDPs
    pdp_urls = [p["pdp_url"] for p in active_products.values()]
    print(f"   🔗 URLs PDPs: {len(pdp_urls)}")
    
    # 6. Procesar PDPs en paralelo
    print(f"\n5️⃣ Procesando PDPs en paralelo...")
    results = process_pdps_parallel(pdp_urls, max_workers=MAX_PARALLEL_WORKERS)
    
    # 7. Actualizar cache y generar output
    print(f"\n6️⃣ Actualizando cache y generando output...")
    df_out = build_template_df()
    processed_count = 0
    error_count = 0
    
    for result in results:
        sku = find_sku_by_url(cache, result["url"])
        
        if not sku or sku not in cache["by_sku"]:
            continue
        
        if result["success"] and result["data"]:
            # ÉXITO: Encontrar matching con StatusBooks
            cache_product = cache["by_sku"][sku]
            stylecolor = cache_product.get("stylecolor")
            
            if stylecolor:
                # Generar variantes para matching
                for variant in normalize_stylecolor_for_match(stylecolor):
                    if variant in sb_map:
                        sb_row = sb_map[variant]
                        
                        # Crear fila de output
                        row = create_output_row(cache_product, result["data"], sb_row)
                        df_out.loc[len(df_out)] = row
                        
                        processed_count += 1
                        break
                else:
                    print(f"   ⚠️  StyleColor {stylecolor} no encontrado en StatusBooks")
            
            # Actualizar cache como exitoso
            update_cache_product(
                cache=cache,
                sku=sku,
                pdp_url=result["url"],
                stylecolor=cache_product.get("stylecolor"),
                is_active=True,
                error_reason=None
            )
            
        else:
            # ERROR: Actualizar cache
            error_count += 1
            
            update_cache_product(
                cache=cache,
                sku=sku,
                pdp_url=result["url"],
                stylecolor=cache["by_sku"][sku].get("stylecolor"),
                is_active=False,
                error_reason=result.get("error", "Unknown error")
            )
    
    # 8. Guardar cache actualizado
    save_json(CACHE_FILE, cache)
    
    # 9. Exportar resultados
    print(f"\n7️⃣ Exportando resultados...")
    
    if df_out.empty:
        print("   ⚠️  No se generaron filas de output")
        print("   💡 Posibles causas:")
        print("      • Productos sin StyleColor en cache")
        print("      • StyleColor no encontrado en StatusBooks")
        print("      • PDPs inactivas o con errores")
        
        return
    
    ts = now_ts()
    csv_path = f"solodeportes_vs_nike_{ts}.csv"
    xlsx_path = f"solodeportes_vs_nike_{ts}.xlsx"
    
    df_out.to_csv(csv_path, index=False, encoding="utf-8-sig")
    
    try:
        export_xlsx(df_out, xlsx_path)
        print(f"   📄 XLSX: {xlsx_path}")
    except Exception as e:
        print(f"   ⚠️  Error exportando XLSX: {e}")
    
    print(f"   📄 CSV:  {csv_path}")
    
    # 10. Resumen final
    print(f"\n" + "=" * 70)
    print("✅ PROCESO COMPLETADO")
    print("=" * 70)
    print(f"📊 Cache:")
    print(f"   • Productos totales: {len(cache.get('by_sku', {}))}")
    print(f"   • Productos activos: {sum(1 for p in cache.get('by_sku', {}).values() if p.get('active', True))}")
    print(f"   • Última actualización: {cache.get('meta', {}).get('last_updated', 'N/A')}")
    print(f"\n📊 Procesamiento:")
    print(f"   • PDPs procesadas: {len(results)}")
    print(f"   • PDPs exitosas: {processed_count}")
    print(f"   • PDPs con error: {error_count}")
    print(f"\n📊 Output:")
    print(f"   • Filas generadas: {len(df_out)}")
    print(f"   • Archivos creados: {csv_path}, {xlsx_path if os.path.exists(xlsx_path) else 'N/A'}")
    print("=" * 70)


if __name__ == "__main__":
    main()