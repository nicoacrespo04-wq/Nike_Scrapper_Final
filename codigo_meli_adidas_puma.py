# -*- coding: utf-8 -*-
"""
MELI MULTI-BRAND SCRAPER - NIKE VS ADIDAS VS PUMA v11
======================================================
Cambios respecto a v10:
- PUMA: nueva marca incorporada
  → Tienda oficial: mercadolibre.com.ar/tienda/puma
  → Excel col 7 = franquicias Puma
  → Verificación de vendedor en PDP incluye "puma"
- EXCEL: nuevo formato de una sola hoja con cols Nike(1)/Adidas(2)/Cat(3)/Puma(7)
  → read_franchises_from_excel() actualizado para leer col 7
- CATEGORÍAS:
  → FUTBOL usa prefijo "botines" (no "zapatillas") en búsquedas generales
  → Sportswear y Running sin prefijo de calzado
- REVOLUTION FIX: fallback de título desde texto visible del item en PLP
  cuando el slug de URL no contiene el nombre de la franquicia
- 3 WORKERS en PLP y PDP (antes 2/3)
- TODAS LAS CATEGORÍAS activas por defecto (CATEGORIA_FILTRO = None)

Cambios anteriores (v10):
- Scroll incremental para lazy loading (fix MLAs faltantes)
- Excluir Salomon en PLP
- Re-scraping de talles == 1 al final del pipeline
- Búsqueda desde dentro de la tienda oficial (v9)
"""

import os
import sys
import json
import time
import re
import uuid
import threading
import random
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
import urllib3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

DEFAULT_EXCEL_PATH = "Comparativa_Nike_Adidas.xlsx"

# Activar/desactivar scraping de Puma (False = solo Nike + Adidas)
SCRAP_PUMA = False

# Proxy SmartProxy
PROXY_HOST = "proxy.smartproxy.net"
PROXY_PORT = 3120
PROXY_USER = "smart-hysjlehcrm30"
PROXY_PASS = "GmpKHg6LdhAbs9Tx"

# Timeouts (ms)
NAV_TIMEOUT   = 60_000
CLICK_TIMEOUT = 15_000   # tiempo máximo esperando que aparezca el dropdown
SCROLL_WAIT   = 1.5

# Límites
MAX_PAGINAS                  = 2
MAX_PRODUCTOS_POR_FRANQUICIA = 0  # 0 = sin límite
MAX_PDP_RETRIES              = 2

# Tope de markdown aceptado para evitar falsos full_price inflados
MAX_MARKDOWN_PCT = 0.60

# Workers
PLP_WORKERS = 3
PDP_WORKERS = 3

# Cache
CACHE_DIR      = "meli_cache"
CACHE_TTL_DAYS = 0   # 0 = sin expiración

# ============================================================
# LOGGING
# ============================================================

def log_info(msg):     print(f"[{datetime.now():%H:%M:%S} INFO ] {msg}")
def log_success(msg):  print(f"[{datetime.now():%H:%M:%S}  ✅  ] {msg}")
def log_warning(msg):  print(f"[{datetime.now():%H:%M:%S}  ⚠️  ] {msg}")
def log_error(msg):    print(f"[{datetime.now():%H:%M:%S}  ❌  ] {msg}")
def log_scraping(msg): print(f"[{datetime.now():%H:%M:%S}  🔍  ] {msg}")
def log_proxy(msg):    print(f"[{datetime.now():%H:%M:%S}  🔌  ] {msg}")
def log_sizes(msg):    print(f"[{datetime.now():%H:%M:%S}  👟  ] {msg}")

# ============================================================
# PROXY
# ============================================================

def build_proxy(session_id: Optional[str] = None) -> Dict[str, str]:
    return {
        "server":   f"http://{PROXY_HOST}:{PROXY_PORT}",
        "username": PROXY_USER,
        "password": PROXY_PASS,
    }


def test_proxy_simple() -> Tuple[bool, Optional[int]]:
    """Test rápido de conectividad del proxy."""
    log_proxy("Testeando proxy...")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                proxy=build_proxy(),
                args=["--ignore-certificate-errors"],
            )
            page = browser.new_page()
            resp = page.goto("https://api.ipify.org?format=json", timeout=15_000)
            if resp and resp.status == 200:
                ip = json.loads(page.inner_text("body")).get("ip", "N/A")
                browser.close()
                log_success(f"Proxy OK — IP: {ip}")
                return True, PROXY_PORT
            browser.close()
    except Exception as e:
        log_warning(f"  Proxy falló: {str(e)[:60]}")
    log_error("No se pudo conectar con el proxy.")
    return False, None

# ============================================================
# LECTURA DE EXCEL
# ============================================================

def read_franchises_from_excel(excel_path: str) -> List[Dict[str, str]]:
    """
    Lee el Excel y devuelve lista de dicts {marca, categoria, franquicia}.

    Formato soportado — una sola hoja 'Competitor' con columnas fijas:
      col 1 = Nike
      col 2 = Adidas
      col 3 = Categoría
      col 7 = Puma

    Fallback formato B: múltiples solapas con headers nike/adidas/puma por columna.
    """
    log_info(f"📖 Leyendo franquicias desde: {excel_path}")
    if not Path(excel_path).exists():
        raise FileNotFoundError(f"No existe el archivo: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    franquicias: List[Dict[str, str]] = []

    # Leer headers de fila 1 para detectar formato
    headers = {str(ws.cell(row=1, column=c).value or "").strip().lower(): c
               for c in range(1, ws.max_column + 1)}

    # ── Formato A: hoja única con col fija Nike(1)/Adidas(2)/Cat(3)/Puma(7) ──
    if "nike" in headers and "adidas" in headers:
        log_info("📋 Formato A: hoja única Nike/Adidas/Puma")

        # Detectar columnas dinámicamente desde los headers
        col_nike   = headers.get("nike",  1)
        col_adidas = headers.get("adidas", 2)
        col_cat    = headers.get("categoría", headers.get("categoria", 3))
        col_puma   = headers.get("puma",  7)

        log_info(f"  Columnas → Nike:{col_nike} Adidas:{col_adidas} "
                 f"Cat:{col_cat} Puma:{col_puma}")

        # Mapa marca → columna
        marca_cols = [
            ("nike",   col_nike),
            ("adidas", col_adidas),
            ("puma",   col_puma),
        ]

        for row in range(2, ws.max_row + 1):
            cat_val = ws.cell(row=row, column=col_cat).value
            if not cat_val or not isinstance(cat_val, str) or not cat_val.strip():
                continue
            categoria = cat_val.strip()

            for marca, col in marca_cols:
                cell = ws.cell(row=row, column=col).value
                if cell and isinstance(cell, str) and cell.strip():
                    franquicias.append({
                        "marca":      marca,
                        "categoria":  categoria,
                        "franquicia": cell.strip(),
                    })

    # ── Formato B: múltiples solapas ──────────────────────────────────────────
    else:
        log_info("📋 Formato B: múltiples solapas por categoría")
        sheet_map = {
            "Botines": "FUTBOL", "Running": "Running",
            "Training": "Training", "Sportswear": "Sportswear",
            "Indumentaria": "Indumentaria", "Accesorios": "Accesorios",
            "FUTBOL": "FUTBOL",
        }
        for sheet_name, category in sheet_map.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws_cat = wb[sheet_name]
            hdrs = [str(ws_cat.cell(row=1, column=c).value or "").strip().lower()
                    for c in range(1, ws_cat.max_column + 1)]
            for row in range(2, ws_cat.max_row + 1):
                for col_idx, marca in enumerate(hdrs, 1):
                    if marca not in ("nike", "adidas", "puma"):
                        continue
                    cell = ws_cat.cell(row=row, column=col_idx).value
                    if cell and str(cell).strip():
                        franquicias.append({"marca": marca, "categoria": category,
                                            "franquicia": str(cell).strip()})

    log_success(f"Total franquicias: {len(franquicias)}")
    marcas_count = defaultdict(int)
    cats_count   = defaultdict(int)
    for f in franquicias:
        marcas_count[f["marca"]] += 1
        cats_count[f["categoria"]] += 1
    for m, n in marcas_count.items():
        log_info(f"  {m.capitalize()}: {n} franquicias")
    for c, n in cats_count.items():
        log_info(f"  {c}: {n} entradas")
    return franquicias

# ============================================================
# PARSER DE PRECIO
# ============================================================

def money_str_to_float(s: str) -> float:
    """Convierte string de precio ARS ('1.234,56' o '1234.56') a float."""
    if not s:
        return 0.0
    raw     = str(s).replace("\xa0", " ").strip()
    cleaned = re.sub(r"[^\d,.]", "", raw)
    if not cleaned:
        return 0.0
    try:
        if "." in cleaned and "," in cleaned:
            return float(cleaned.replace(".", "").replace(",", "."))
        if "." in cleaned and re.search(r"\d{1,3}(?:\.\d{3})+$", cleaned):
            return float(cleaned.replace(".", ""))
        if "." in cleaned:
            return float(cleaned)
        if "," in cleaned:
            return float(cleaned.replace(",", "."))
        return float(cleaned)
    except Exception:
        return 0.0

# ============================================================
# EXTRACCIÓN DE PRECIOS
# ============================================================

def extract_prices(page) -> Tuple[Optional[float], Optional[float], Optional[str], Optional[str]]:
    """
    Devuelve (full_price, final_price, full_raw, final_raw).

    Estrategia robusta:
      final_price → meta[itemprop=price]  (fuente más confiable, siempre el precio de compra)
      full_price  → precio tachado (.ui-pdp-price__original-value) si existe,
                    sino el primer .andes-money-amount__fraction del bloque de precio principal
    """
    full_price = final_price = None
    full_raw   = final_raw   = None

    # ── Final price: siempre desde meta (el precio real de compra) ───────────
    try:
        meta = page.locator('meta[itemprop="price"]').first
        if meta.count() > 0:
            final_raw   = meta.get_attribute("content")
            final_price = money_str_to_float(final_raw)
            log_info(f"    Final price (meta): ${final_price:,.0f}")
    except Exception:
        pass

    # Fallback final price: segunda línea de precio
    if not final_price:
        try:
            second = page.locator(
                "span.ui-pdp-price__second-line .andes-money-amount__fraction"
            ).first
            if second.count() > 0:
                final_raw   = second.inner_text()
                final_price = money_str_to_float(final_raw)
                log_info(f"    Final price (second line): ${final_price:,.0f}")
        except Exception:
            pass

    # ── Full price: buscar precio tachado (precio original sin descuento) ────
    # Se recolectan candidatos y se elige el MENOR >= final para evitar
    # falsos positivos inflados (ej: valores cruzados de otros bloques).
    original_selectors = [
        "span.ui-pdp-price__original-value .andes-money-amount__fraction",
        ".ui-pdp-price__original-value .andes-money-amount__fraction",
        "s .andes-money-amount__fraction",                     # tachado genérico
        ".andes-money-amount--previous .andes-money-amount__fraction",
        "[class*='original'] .andes-money-amount__fraction",
        "[class*='strike'] .andes-money-amount__fraction",
    ]

    candidatos_full: List[Tuple[float, str, str]] = []
    for sel in original_selectors:
        try:
            loc = page.locator(sel)
            n = min(loc.count(), 5)
            for i in range(n):
                txt = loc.nth(i).inner_text()
                val = money_str_to_float(txt)
                if val and val > 0:
                    candidatos_full.append((val, txt, sel))
        except Exception:
            continue

    if candidatos_full:
        if final_price and final_price > 0:
            validos = [c for c in candidatos_full if c[0] >= final_price]
            elegido = min(validos, key=lambda x: x[0]) if validos else min(candidatos_full, key=lambda x: x[0])
        else:
            elegido = min(candidatos_full, key=lambda x: x[0])
        full_price, full_raw, sel_usado = elegido
        log_info(f"    Full price (tachado [{sel_usado}]): ${full_price:,.0f}")

    # Si no hay precio tachado → full price = final price (sin descuento)
    if not full_price and final_price:
        full_price = final_price
        full_raw   = final_raw
        log_info(f"    Full price = Final price (sin descuento): ${full_price:,.0f}")

    # Sanity extra: evita falsos descuentos extremos por lectura errónea
    if full_price and final_price and full_price > 0:
        markdown_pct = (full_price - final_price) / full_price
        if markdown_pct < 0 or markdown_pct > MAX_MARKDOWN_PCT:
            log_warning(
                f"    ⚠️  markdown sospechoso ({markdown_pct:.1%}) "
                f"full=${full_price:,.0f} final=${final_price:,.0f} — usando full=final"
            )
            full_price = final_price
            full_raw = final_raw

    return full_price, final_price, full_raw, final_raw

# ============================================================
# EXTRACCIÓN DE CUOTAS
# ============================================================

def extract_installments(page, final_price: float) -> int:
    """
    Busca 'X cuotas de $YYY' en el body y valida que X*YYY ≈ final_price (±2%).
    Devuelve la mayor cantidad válida encontrada.
    """
    if not final_price or final_price <= 0:
        return 0
    try:
        text = page.inner_text("body")
    except Exception:
        return 0

    mejor = 0
    patrones = [
        r"(\d{1,2})\s*cuotas?\s*de\s*\$?\s*([\d\.]+(?:,\d{1,2})?)",
        r"mismo\s*precio\s*en\s*(\d{1,2})\s*cuotas?\s*de\s*\$?\s*([\d\.]+(?:,\d{1,2})?)",
    ]
    for patron in patrones:
        for m in re.finditer(patron, text, re.IGNORECASE):
            try:
                cuotas = int(m.group(1))
                monto  = money_str_to_float(m.group(2))
                if cuotas <= 0 or monto <= 0:
                    continue
                diff = abs(cuotas * monto - final_price) / final_price
                if diff <= 0.02 and cuotas > mejor:
                    mejor = cuotas
                    log_info(f"    ✓ {cuotas} cuotas (diff {diff:.1%})")
            except Exception:
                continue
    return mejor

# ============================================================
# EXTRACCIÓN DE TALLES — v8: SELECTORES EXACTOS DEL DOM REAL
# ============================================================
#
# Confirmado por diagnóstico sobre MLA-2583781938:
#
#   CONTENEDOR : [data-testid='PICKER-SIZE']
#                class='ui-pdp-outside_variations__picker'
#
#   TRIGGER    : button.andes-dropdown__trigger  (DENTRO del contenedor)
#                role=combobox, aria-haspopup=listbox, aria-expanded=false→true
#                ⚠️  También existe en "Cantidad" → NUNCA buscar en página entera
#
#   ITEMS      : [role='listbox'] [role='option']   (tras abrir el dropdown)
#                class='andes-list__item andes-list__item--size-medium'
#                texto directo en inner_text: '10 UK', '10.5 UK', etc.
#
#   TALLES VISIBLES (sin dropdown, botones directos):
#                [data-testid='PICKER-SIZE'] a[role='button']
#                class='ui-pdp-outside_variations__thumbnails__item'
#                aria-label='Botón N de M, ...'
#
# ─────────────────────────────────────────────────────────────

# Palabras que indican que un item NO es un talle válido
_TEXTOS_IGNORAR = {
    "elegí", "elegir", "seleccionar", "seleccioná", "ver más",
    "cantidad", "unidad", "unidades", "",
}


def _is_disabled(item) -> bool:
    """True si el elemento está marcado como sin stock / deshabilitado."""
    try:
        if item.get_attribute("disabled"):
            return True
        if item.get_attribute("aria-disabled") == "true":
            return True
        cls = item.get_attribute("class") or ""
        if any(k in cls for k in ("disabled", "out-of-stock", "no-stock",
                                  "unavailable", "selected--disabled")):
            return True
    except Exception:
        pass
    return False


def _text_es_talle(texto: str) -> bool:
    """True si el texto parece un talle real (número, letra, UK, US, etc.)."""
    if not texto:
        return False
    if texto.lower() in _TEXTOS_IGNORAR:
        return False
    # Debe contener al menos un dígito o una letra de talle (XS/S/M/L/XL/XXL)
    if re.search(r"\d", texto):
        return True
    if re.match(r"^(XXS|XS|S|M|L|XL|XXL|XXXL|XG|GG|EG)$", texto.strip(), re.IGNORECASE):
        return True
    return False


def _leer_items_listbox(page) -> int:
    """
    Lee los items de un [role='listbox'] ya abierto.
    El popper/listbox en Meli se renderiza en el body via Tippy.js,
    NO dentro del contenedor PICKER-SIZE. Hay que buscarlo en toda la página.
    Confirmado por diagnóstico: el listbox tiene id terminado en -menu-list
    y los items tienen role=option con clase andes-list__item.
    """
    candidatos = [
        # Más específico: listbox del dropdown de talles (NO de cantidad)
        "[data-testid='PICKER-SIZE'] ~ * [role='option']",  # hermano
        "[data-testid='popper'] [role='option']",            # popper de Tippy
        "[data-testid='popper'] li",
        "[role='listbox'][aria-label*='Talle'] [role='option']",
        "[role='listbox'][aria-label*='talle'] [role='option']",
        "[role='listbox'][aria-label*='Size'] [role='option']",
        # Genérico pero funciona si hay un solo listbox abierto
        "[role='listbox'] [role='option']",
        "[role='listbox'] li",
        ".andes-list__item",
    ]
    for sel in candidatos:
        try:
            items = page.locator(sel).all()
            if not items:
                continue
            disponibles = []
            for item in items:
                if _is_disabled(item):
                    continue
                try:
                    texto = item.inner_text().strip()
                except Exception:
                    texto = ""
                if _text_es_talle(texto):
                    disponibles.append(texto)
            if disponibles:
                log_sizes(f"    {len(disponibles)} talles en listbox [{sel}]: "
                          f"{chr(44).join(disponibles[:6])}{'…' if len(disponibles) > 6 else ''}")
                return len(disponibles)
        except Exception:
            continue
    return 0


def _leer_items_visibles(container) -> int:
    """
    Lee talles que ya están visibles como botones/thumbnails (sin dropdown).
    Usado cuando el picker muestra chips directamente en pantalla.
    """
    # Scroll horizontal por si hay overflow
    try:
        container.evaluate("el => el.scrollLeft = el.scrollWidth")
        time.sleep(0.4)
    except Exception:
        pass

    # Selectores de items visibles dentro del contenedor
    candidatos = [
        "a[role='button']",          # thumbnails con href
        "button:not(.andes-dropdown__trigger)",  # botones que no sean el trigger
        "[role='button']",
        "li",
        ".ui-pdp-outside_variations__thumbnails__item",
        ".ui-pdp-variations__thumbnail",
    ]
    for sel in candidatos:
        try:
            items = container.locator(sel).all()
            if not items:
                continue
            disponibles = []
            for item in items:
                if _is_disabled(item):
                    continue
                # Intentar texto directo
                try:
                    texto = item.inner_text().strip()
                except Exception:
                    texto = ""
                # Fallback: aria-label (ej: "Botón 1 de 9, 9, 41.5 AR")
                if not _text_es_talle(texto):
                    try:
                        aria = item.get_attribute("aria-label") or ""
                        # Extraer el valor de talle del aria-label
                        m = re.search(
                            r"(?:Botón\s+\d+\s+de\s+\d+,\s*\d+,\s*)(.+?)(?:\s+AR|\s+EU|\s+US|\s+UK|$)",
                            aria, re.IGNORECASE
                        )
                        if m:
                            texto = m.group(1).strip()
                        else:
                            # Fallback más simple: cualquier número con decimales
                            m2 = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:AR|EU|US|UK|BR|CM)?", aria)
                            if m2:
                                texto = m2.group(1)
                    except Exception:
                        pass
                if _text_es_talle(texto):
                    disponibles.append(texto)
            if disponibles:
                log_sizes(f"    {len(disponibles)} talles visibles [{sel}]: "
                          f"{', '.join(disponibles[:6])}{'…' if len(disponibles) > 6 else ''}")
                return len(disponibles)
        except Exception:
            continue
    return 0


# JavaScript para leer listbox de talles (Tippy popper, dinámico)
_JS_LEER_LISTBOX = """
(function() {
    var listboxes = Array.from(document.querySelectorAll('[role="listbox"]'));
    for (var i = 0; i < listboxes.length; i++) {
        var lb = listboxes[i];
        var label = (lb.getAttribute('aria-label') || '').toLowerCase();
        if (label.indexOf('cantidad') >= 0 || label.indexOf('quantity') >= 0) continue;
        var options = Array.from(lb.querySelectorAll('[role="option"]'));
        if (options.length === 0) continue;
        var disponibles = [];
        for (var j = 0; j < options.length; j++) {
            var opt = options[j];
            if (opt.getAttribute('aria-disabled') === 'true') continue;
            if (opt.className.indexOf('disabled') >= 0) continue;
            var text = (opt.textContent || '').trim();
            if (!text) continue;
            if (/[0-9]/.test(text) || /^(XS|S|M|L|XL|XXL|XXXL|XG|GG|EG)$/i.test(text)) {
                disponibles.push(text);
            }
        }
        if (disponibles.length > 0) {
            return {count: disponibles.length, talles: disponibles.slice(0, 8)};
        }
    }
    return null;
})()
"""

# JavaScript para leer talles visibles (chips/thumbnails)
_JS_LEER_VISIBLES = """
(function() {
    var picker = document.querySelector('[data-testid="PICKER-SIZE"]');
    if (!picker) return null;
    var items = Array.from(picker.querySelectorAll('a, button'));
    var disponibles = [];
    var ignorar = ['elegi', 'elegir', 'seleccionar', 'selecciona'];
    for (var i = 0; i < items.length; i++) {
        var item = items[i];
        if (item.hasAttribute('disabled')) continue;
        if (item.getAttribute('aria-disabled') === 'true') continue;
        if (item.className.indexOf('andes-dropdown__trigger') >= 0) continue;
        var text = (item.textContent || '').trim();
        if (!text || ignorar.indexOf(text.toLowerCase()) >= 0) {
            var aria = item.getAttribute('aria-label') || '';
            var m = aria.match(/([0-9]+(?:[.,][0-9]+)?)/);
            if (m) text = m[1];
        }
        if (!text) continue;
        if (/[0-9]/.test(text) || /^(XS|S|M|L|XL|XXL|XXXL|XG|GG|EG)$/i.test(text.trim())) {
            if (ignorar.indexOf(text.toLowerCase()) < 0) {
                disponibles.push(text);
            }
        }
    }
    return disponibles.length > 0 ? {count: disponibles.length, talles: disponibles.slice(0, 8)} : null;
})()
"""

# JavaScript para hacer click en el trigger de talles
_JS_CLICK_TRIGGER = """
(function() {
    var picker = document.querySelector('[data-testid="PICKER-SIZE"]');
    if (!picker) return false;
    var trigger = picker.querySelector('button.andes-dropdown__trigger');
    if (!trigger) return false;
    trigger.click();
    return true;
})()
"""

# JavaScript para verificar si el listbox de talles está abierto
_JS_LISTBOX_ABIERTO = """
(function() {
    var lbs = document.querySelectorAll('[role="listbox"]');
    for (var i = 0; i < lbs.length; i++) {
        var label = (lbs[i].getAttribute('aria-label') || '').toLowerCase();
        if (label.indexOf('cantidad') >= 0) continue;
        if (lbs[i].querySelectorAll('[role="option"]').length > 0) return true;
    }
    return false;
})()
"""


def _leer_listbox_via_js(page) -> int:
    """Lee el listbox de talles abierto via JavaScript."""
    try:
        resultado = page.evaluate(_JS_LEER_LISTBOX)
        if resultado and resultado.get("count", 0) > 0:
            talles = resultado.get("talles", [])
            log_sizes(f"    {resultado['count']} talles via JS: {', '.join(talles)}{'…' if resultado['count'] > 8 else ''}")
            return resultado["count"]
    except Exception as e:
        log_sizes(f"    JS listbox error: {e}")
    return 0


def _leer_visibles_via_js(page) -> int:
    """Lee talles visibles (chips) via JavaScript."""
    try:
        resultado = page.evaluate(_JS_LEER_VISIBLES)
        if resultado and resultado.get("count", 0) > 0:
            talles = resultado.get("talles", [])
            log_sizes(f"    {resultado['count']} talles visibles JS: {', '.join(talles)}{'…' if resultado['count'] > 8 else ''}")
            return resultado["count"]
    except Exception as e:
        log_sizes(f"    JS visibles error: {e}")
    return 0


def extract_sizes(page) -> Optional[int]:
    """
    Extrae la cantidad de talles disponibles de una PDP de Meli.

    Retorna:
      int  → cantidad real de talles (incluyendo 1 si solo hay 1 talle visible)
      None → se detectó dropdown de talles pero no se pudo leer su contenido
             (no rompe promedios en el output)
      0    → no se encontró picker de talles en absoluto

    Estrategia:
    ───────────────────────────────────────────────────────
    1. Scroll para montar componentes lazy.
    2. Buscar [data-testid='PICKER-SIZE'].
    3a. Si tiene button.andes-dropdown__trigger → modo DROPDOWN
        → click + wait listbox → leer opciones
        → si falla tras retry → None (dropdown detectado pero ilegible)
    3b. Si no tiene trigger → modo VISIBLE
        → leer chips/thumbnails → count real (puede ser 1, es válido)
    """
    try:
        # ── 1. Scroll para montar lazy components ────────────────────────────
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(0.8)
        page.evaluate("window.scrollTo(0, 0)")
        time.sleep(0.5)

        picker = page.locator("[data-testid='PICKER-SIZE']").first
        if picker.count() == 0:
            log_sizes("    Sin PICKER-SIZE — buscando fallback...")
            return _extract_sizes_fallback(page)

        log_sizes("    Contenedor: [data-testid='PICKER-SIZE']")

        trigger = picker.locator("button.andes-dropdown__trigger").first
        if trigger.count() > 0:
            # ── 3a. DROPDOWN ─────────────────────────────────────────────────
            log_sizes("    Modo: DROPDOWN")
            try:
                try:
                    page.wait_for_selector(
                        "[data-testid='PICKER-SIZE'] button.andes-dropdown__trigger",
                        state="visible", timeout=8000
                    )
                except Exception:
                    pass
                time.sleep(0.5)

                trigger.click(timeout=CLICK_TIMEOUT)

                listbox_sel = (
                    "[role='listbox'][aria-label*='Talle'], "
                    "[role='listbox'][aria-label*='talle'], "
                    "[role='listbox'][aria-label*='Size']"
                )
                try:
                    page.wait_for_selector(listbox_sel, timeout=10000)
                except Exception:
                    time.sleep(3.0)

                count = _leer_listbox_playwright(page)
                if count > 0:
                    return count

                # Dropdown detectado pero ilegible → None (se excluye del promedio)
                log_sizes("    Dropdown ilegible → None")
                return None

            except Exception as e:
                log_sizes(f"    Error click trigger: {e}")
                # Si ni siquiera se pudo clickear el dropdown → None
                return None

        # ── 3b. VISIBLE: chips/thumbnails directos ───────────────────────────
        # En este modo, talle 1 es válido (producto genuinamente en un solo talle)
        log_sizes("    Modo: VISIBLE")
        count = _leer_visibles_via_js(page)
        if count > 0:
            return count
        count = _leer_items_visibles(picker)
        if count > 0:
            return count

        log_sizes("    Sin talles encontrados")
        return 0

    except Exception as e:
        log_sizes(f"    Error en extract_sizes: {e}")
        return 0


def _leer_listbox_playwright(page) -> int:
    """
    Lee las opciones del listbox de talles ya abierto.
    Busca el listbox con aria-label que contenga 'Talle' o 'Size',
    excluyendo el de 'Cantidad'.
    Estrategia confirmada: [role='listbox'] con aria-label 'Talle: Elegí'
    """
    # Selectores en orden de especificidad
    selectores_listbox = [
        "[role='listbox'][aria-label*='Talle'] [role='option']",
        "[role='listbox'][aria-label*='talle'] [role='option']",
        "[role='listbox'][aria-label*='Size'] [role='option']",
        "[role='listbox'][aria-label*='size'] [role='option']",
    ]
    for sel in selectores_listbox:
        try:
            items = page.locator(sel).all()
            if not items:
                continue
            disponibles = []
            for item in items:
                if _is_disabled(item):
                    continue
                try:
                    texto = item.inner_text().strip()
                except Exception:
                    texto = ""
                if _text_es_talle(texto):
                    disponibles.append(texto)
            if disponibles:
                log_sizes(f"    {len(disponibles)} talles [{sel[:40]}]: "
                          f"{', '.join(disponibles[:6])}{'…' if len(disponibles) > 6 else ''}")
                return len(disponibles)
        except Exception:
            continue

    # Fallback: cualquier listbox que no sea de cantidad
    try:
        lbs = page.locator("[role='listbox']").all()
        for lb in lbs:
            try:
                label = (lb.get_attribute("aria-label") or "").lower()
                if "cantidad" in label or "quantity" in label:
                    continue
                items = lb.locator("[role='option']").all()
                disponibles = []
                for item in items:
                    if _is_disabled(item):
                        continue
                    try:
                        texto = item.inner_text().strip()
                    except Exception:
                        texto = ""
                    if _text_es_talle(texto):
                        disponibles.append(texto)
                if disponibles:
                    log_sizes(f"    {len(disponibles)} talles [fallback listbox label='{label}']: "
                              f"{', '.join(disponibles[:6])}{'…' if len(disponibles) > 6 else ''}")
                    return len(disponibles)
            except Exception:
                continue
    except Exception:
        pass

    return 0


def _extract_sizes_fallback(page) -> int:
    """Fallback para PDPs sin [data-testid='PICKER-SIZE']."""
    for sel in [".ui-pdp-outside_variations__picker", ".ui-pdp-variations__picker"]:
        try:
            container = page.locator(sel).first
            if container.count() == 0:
                continue
            log_sizes(f"    Fallback contenedor: {sel}")
            trigger = container.locator("button.andes-dropdown__trigger").first
            if trigger.count() > 0:
                trigger.click(timeout=CLICK_TIMEOUT)
                time.sleep(1.5)
                count = _leer_listbox_playwright(page)
                if count > 0:
                    return count
            count = _leer_items_visibles(container)
            if count > 0:
                return count
        except Exception:
            continue
    return 0


def setup_route_blocking(context):
    """Bloquea imágenes, fuentes y media. Permite JS y CSS."""
    def handler(route):
        try:
            if route.request.resource_type in ("image", "font", "media"):
                route.abort()
            else:
                route.continue_()
        except Exception:
            try:
                route.continue_()
            except Exception:
                pass
    context.route("**/*", handler)
    return context

# ============================================================
# CONSTRUCCIÓN DE URLs
# ============================================================

# Categorías que usan "zapatillas" como prefijo de búsqueda
CATEGORIAS_ZAPATILLAS = {"running", "training", "sportswear"}

# Categorías que usan "botines" como prefijo de búsqueda
CATEGORIAS_BOTINES = {"futbol", "fútbol"}

# Todas las categorías de calzado (para compatibilidad con código legado)
CATEGORIAS_CALZADO = CATEGORIAS_ZAPATILLAS | CATEGORIAS_BOTINES


def _prefijo_categoria(categoria: str) -> str:
    """Devuelve el prefijo de calzado correcto según la categoría."""
    cat_lower = categoria.strip().lower()
    if cat_lower in CATEGORIAS_BOTINES:
        return "botines"
    if cat_lower in CATEGORIAS_ZAPATILLAS:
        return "zapatillas"
    return ""  # Sportswear puro, indumentaria, etc. → sin prefijo

# ============================================================
# PARSING DE FRANQUICIAS CON COMILLAS
# ============================================================

def parsear_franquicia(nombre_raw: str) -> dict:
    """
    Parsea el nombre de la franquicia del Excel y genera reglas de matching.

    Sintaxis soportada:
      "ZOOM FLY"          → frase_exacta: "zoom fly" debe aparecer JUNTA en el título
      PEGASUS "PLUS"      → "plus" obligatoria, "pegasus" libre (al menos 1)
      ADIZERO "BOSTON"    → "boston" obligatoria, "adizero" libre
      SUPERNOVA "STRIDE"  → "stride" obligatoria, "supernova" libre
      VAPORFLY            → "vaporfly" libre
      ADIZERO ADIOS PRO   → al menos 2 de ["adizero","adios","pro"] en título

    Devuelve:
      nombre_limpio        str   — sin comillas, para display y búsqueda URL
      frase_exacta         str|None — si TODO está entre comillas → substring exacto
      palabras_libres      list  — palabras sin comillas (al menos 1 debe aparecer)
      palabras_obligatorias list — palabras entre comillas (TODAS deben aparecer)
    """
    nombre = nombre_raw.strip()
    # Caso 1: TODO entre comillas → "ZOOM FLY"
    m_todo = re.match(r'^"([^"]+)"$', nombre)
    if m_todo:
        frase = m_todo.group(1).strip().lower()
        return {
            "nombre_limpio": m_todo.group(1).strip(),
            "frase_exacta": frase,
            "palabras_libres": [],
            "palabras_obligatorias": [],
        }
    # Caso 2: mezcla de palabras libres y entre comillas
    palabras_libres      = []
    palabras_obligatorias = []
    tokens = re.findall(r'\"([^\"]+)\"|(\S+)', nombre)
    for quoted, libre in tokens:
        if quoted:
            for w in quoted.strip().lower().split():
                if len(w) > 1:
                    palabras_obligatorias.append(w)
        elif libre:
            w = libre.strip().lower()
            if len(w) > 1:
                palabras_libres.append(w)
    nombre_limpio = re.sub(r'"', '', nombre).strip()
    return {
        "nombre_limpio": nombre_limpio,
        "frase_exacta": None,
        "palabras_libres": palabras_libres,
        "palabras_obligatorias": palabras_obligatorias,
    }


def match_franquicia(titulo: str, reglas: dict) -> bool:
    """
    Verifica si el título (extraído de la URL de MLA) matchea las reglas de la franquicia.
    Todo normalizado a minúsculas.
    """
    titulo_lower = titulo.lower()

    # Frase exacta: deben estar juntas y seguidas
    if reglas["frase_exacta"]:
        return reglas["frase_exacta"] in titulo_lower

    # Palabras obligatorias (entre comillas): TODAS deben aparecer
    for palabra in reglas["palabras_obligatorias"]:
        if palabra not in titulo_lower:
            return False

    # Palabras libres: al menos 1 debe aparecer (si las hay)
    if reglas["palabras_libres"]:
        if not any(w in titulo_lower for w in reglas["palabras_libres"]):
            return False

    return True


def build_search_url(marca: str, categoria: str, franquicia: str, page: int = 1) -> str:
    """
    Construye la URL de búsqueda en Mercado Libre Argentina (fallback cuando
    la tienda oficial no tiene el producto).
    Prefijos por categoría:
      Running/Training/Sportswear → "zapatillas"
      FUTBOL                      → "botines"
      Otros                       → sin prefijo
    """
    nombre_limpio = parsear_franquicia(franquicia)["nombre_limpio"]
    prefijo = _prefijo_categoria(categoria)
    if prefijo:
        query = f"{prefijo}-{nombre_limpio}-{marca.capitalize()}".replace(" ", "-")
    else:
        query = f"{nombre_limpio}-{marca.capitalize()}".replace(" ", "-")

    if page == 1:
        return f"https://listado.mercadolibre.com.ar/{query}"
    offset = (page - 1) * 48 + 1
    return f"https://listado.mercadolibre.com.ar/{query}_Desde_{offset}_NoIndex_True"


def build_search_urls_extra(marca: str, categoria: str, franquicia: str) -> List[str]:
    """
    Genera URLs de búsqueda alternativas cuando la query principal trae pocos resultados.
    Ej: VOMERO → también busca "zoom vomero nike", "air vomero nike"
    """
    nombre_limpio = parsear_franquicia(franquicia)["nombre_limpio"]
    cat_lower     = categoria.strip().lower()
    marca_cap     = marca.capitalize()
    franq_slug    = nombre_limpio.replace(" ", "-")
    urls = []
    if cat_lower in CATEGORIAS_CALZADO:
        urls.append(f"https://listado.mercadolibre.com.ar/{franq_slug}-{marca_cap}")
        urls.append(f"https://listado.mercadolibre.com.ar/zoom-{franq_slug}-{marca_cap}")
        urls.append(f"https://listado.mercadolibre.com.ar/air-{franq_slug}-{marca_cap}")
    return urls


def search_within_store(page, marca: str, franquicia: str, categoria: str = "") -> bool:
    """
    Navega a la tienda oficial de la marca y usa el buscador interno de la página
    para buscar la franquicia. Así Meli mantiene el contexto de tienda oficial
    y todos los resultados son de esa tienda (sin usados, sin revendedores).

    Para categoría FUTBOL antepone "botines" al query para evitar que aparezcan
    zapatillas con el mismo nombre (ej: "Club" trae zapatillas si no se filtra).

    Pasos:
      1. Ir a la página de la tienda oficial
      2. Localizar el input de búsqueda dentro de la tienda
      3. Tipear el query y hacer Enter
      4. Esperar a que carguen los resultados

    Devuelve True si la búsqueda se realizó con éxito, False si algo falló.
    """
    nombre_limpio = parsear_franquicia(franquicia)["nombre_limpio"]

    # Para fútbol, buscar "botines <franquicia>" para evitar falsos positivos
    # con zapatillas que comparten nombre (ej: Adidas Club, Puma Play)
    prefijo = _prefijo_categoria(categoria)
    query   = f"{prefijo} {nombre_limpio}".strip() if prefijo else nombre_limpio

    store_url = TIENDAS_URL_BASE.get(marca.lower())
    if not store_url:
        log_warning(f"No hay URL de tienda para '{marca}'")
        return False

    log_scraping(f"  🏪 Entrando a tienda oficial {marca.upper()} → {store_url}")
    resp = page.goto(store_url, wait_until="domcontentloaded")
    if resp and resp.status >= 400:
        log_warning(f"    Status {resp.status} al cargar la tienda")
        return False

    time.sleep(2)

    # Selectores del buscador interno de la tienda Meli
    # (puede variar; probamos en orden de especificidad)
    search_input_selectors = [
        "input[data-testid='search-box']",
        "input[name='q']",
        "input[placeholder*='buscar' i]",
        "input[placeholder*='Buscar' i]",
        ".nav-search-input",
        "#search-box",
        "input[type='search']",
        "input[type='text'][class*='search']",
    ]

    input_el = None
    for sel in search_input_selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0 and loc.is_visible():
                input_el = loc
                log_info(f"    🔍 Input de búsqueda encontrado: [{sel}]")
                break
        except Exception:
            continue

    if input_el is None:
        log_warning(f"    ⚠️  No se encontró el input de búsqueda en la tienda {marca}")
        return False

    # Limpiar y tipear el query
    try:
        input_el.click()
        time.sleep(0.3)
        input_el.fill("")
        input_el.type(query, delay=60)  # delay humano
        time.sleep(0.5)
        input_el.press("Enter")
        log_info(f"    ⌨️  Buscando '{query}' dentro de la tienda...")
    except Exception as e:
        log_warning(f"    ⚠️  Error al tipear en el buscador: {e}")
        return False

    # Esperar a que carguen los resultados
    try:
        page.wait_for_load_state("domcontentloaded", timeout=NAV_TIMEOUT)
        time.sleep(2)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1)
        page.evaluate("window.scrollTo(0, 0)")
        time.sleep(0.5)
    except Exception as e:
        log_warning(f"    ⚠️  Timeout esperando resultados: {e}")
        # Continuamos igual, puede haber resultados parciales

    log_success(f"    ✅ Búsqueda realizada dentro de la tienda {marca.upper()}")
    return True


def extract_mla_from_url(url: str) -> Optional[str]:
    m = re.search(r"MLA[_-]?(\d+)", url)
    return m.group(1) if m else None


def extract_title_from_url(url: str) -> str:
    try:
        for part in url.split("/"):
            if "MLA-" in part:
                title = re.sub(r"^MLA-\d+-", "", part)
                title = re.sub(r"_[A-Z]+$", "", title)
                return title.replace("-", " ").lower()
    except Exception:
        pass
    return ""

# ============================================================
# DETECCIÓN DE TIENDA OFICIAL
# ============================================================

# Nombres conocidos de tiendas oficiales Nike y Adidas en Meli AR
TIENDAS_OFICIALES = {
    "nike":  ["nike", "tienda nike", "nike argentina", "nike official"],
    "adidas":["adidas", "tienda adidas", "adidas argentina", "adidas official"],
    "puma":  ["puma", "tienda puma", "puma argentina", "puma official"],
}

# URLs base de las tiendas oficiales en Meli AR
# Búsqueda dentro de la tienda: /tienda/{marca}/buscar?q={query}
TIENDAS_URL_BASE = {
    "nike":   "https://www.mercadolibre.com.ar/tienda/nike",
    "adidas": "https://www.mercadolibre.com.ar/tienda/adidas",
    "puma":   "https://www.mercadolibre.com.ar/tienda/puma",
}

def _es_tienda_oficial(item, marca: str) -> bool:
    """
    Verifica si un item de la PLP pertenece a la tienda oficial de la marca.
    Estrategias en orden:
      1. Buscar el badge/label de 'Tienda oficial' en el item
      2. Buscar el nombre del vendedor en el item y comparar con tiendas conocidas
      3. Verificar atributo data-* de tienda oficial
    Si no puede determinarlo, devuelve True (no descartar por las dudas).
    """
    try:
        # Estrategia 1: badge de tienda oficial en el HTML del item
        item_html = item.inner_html().lower()

        # Meli marca las tiendas oficiales con estos textos/clases
        oficial_signals = [
            "tienda oficial",
            "official store",
            "ui-search-official-store",
            "official-store",
            "data-official-store",
        ]
        for signal in oficial_signals:
            if signal in item_html:
                # Adicionalmente verificar que sea la marca correcta
                for nombre in TIENDAS_OFICIALES.get(marca.lower(), []):
                    if nombre in item_html:
                        log_sizes(f"      ✅ Tienda oficial: {nombre}")
                        return True
                # Hay badge oficial pero no matchea la marca → descartar
                log_sizes(f"      ✗  Tienda oficial de OTRA marca")
                return False

        # Estrategia 2: buscar nombre del vendedor
        vendedor_selectors = [
            ".ui-search-official-store-label",
            ".ui-search-item__store-logo-label",
            "[class*='official-store']",
            "[class*='store-label']",
            ".poly-component__seller",
            "[class*='seller']",
        ]
        for sel in vendedor_selectors:
            try:
                el = item.locator(sel).first
                if el.count() > 0:
                    vendedor_text = el.inner_text().lower().strip()
                    for nombre in TIENDAS_OFICIALES.get(marca.lower(), []):
                        if nombre in vendedor_text:
                            log_sizes(f"      ✅ Vendedor oficial: {vendedor_text}")
                            return True
                    # Si encontró vendedor pero no es la marca correcta
                    if vendedor_text:
                        log_sizes(f"      ✗  Vendedor: '{vendedor_text}' no es tienda oficial")
                        return False
            except Exception:
                continue

        # No se pudo determinar → aceptar (para no filtrar demasiado)
        return True

    except Exception as e:
        log_sizes(f"      ⚠️  No se pudo verificar tienda: {e}")
        return True   # en caso de error, no descartar


# ============================================================
# SCRAPEO DE PLP
# ============================================================

def _scroll_hasta_cargar_todo(page, selector: str, max_intentos: int = 15) -> None:
    """
    Hace scroll incremental hasta que no aparezcan items nuevos.
    Necesario porque Meli usa lazy loading: los items se renderizan
    a medida que el viewport los alcanza, no todos de una.

    Estrategia:
      - Scroll en pasos de 600px con pausa para que React monte los nuevos items
      - Se detiene cuando dos pasadas consecutivas devuelven el mismo count
      - Tope de max_intentos para no quedar en loop infinito
    """
    prev_count = 0
    sin_cambio = 0
    for intento in range(max_intentos):
        page.evaluate("window.scrollBy(0, 600)")
        time.sleep(0.8)
        count = page.locator(selector).count()
        log_info(f"    Scroll {intento+1}: {count} items cargados")
        if count == prev_count:
            sin_cambio += 1
            if sin_cambio >= 3:   # 3 pasadas sin cambio = llegamos al final
                break
        else:
            sin_cambio = 0
        prev_count = count
    # Volver arriba para que los primeros items sigan en DOM
    page.evaluate("window.scrollTo(0, 0)")
    time.sleep(0.5)


def scrape_plp_for_franchise(marca: str, categoria: str, franquicia: str) -> List[Dict[str, str]]:
    """
    Scrapea el listado de Meli buscando la franquicia en los resultados de la marca.

    Estrategia v12 — URL directa al listado general:
      1. Navega directo a listado.mercadolibre.com.ar/{prefijo}-{franquicia}-{marca}
         → ~3x más rápido que tipear en buscador interno
         → Los primeros resultados son siempre tienda oficial (verificado en debug)
         → El filtro de vendedor en scrape_pdp() actúa como red de seguridad
      2. Scroll incremental hasta cargar TODOS los items (lazy loading)
      3. Filtra: descarta "salomon" en URL + franquicia en título
    """
    resultados:   List[Dict[str, str]] = []
    mlas_vistos:  set = set()
    desc_franq    = 0
    desc_salomon  = 0
    desc_trail    = 0
    desc_cross    = 0

    MARCAS_EXCLUIR   = ["salomon", "salom"]
    cat_upper        = categoria.strip().upper()
    cat_lower        = categoria.strip().lower()
    # ¿La franquicia misma dice "trail"? Si es así, no filtramos trail.
    franq_es_trail   = "trail" in franquicia.lower()
    # Palabras de accesorio a excluir en categorías de calzado
    PALABRAS_ACCESORIOS = ["mochila", "bolso", "bolsa", "gorra", "remera",
                           "campera", "pantalon", "medias", "calcetines",
                           "guante", "chaleco", "riñonera"]

    # Construir URL principal (pág 1) y páginas adicionales
    nombre_limpio = parsear_franquicia(franquicia)["nombre_limpio"]

    # Páginas 1..MAX_PAGINAS de la query principal
    urls_paginadas = [build_search_url(marca, categoria, franquicia, pg)
                      for pg in range(1, MAX_PAGINAS + 1)]
    # Fallback sin prefijo (ayuda a Club, Revolution, etc.)
    slug_simple    = f"{nombre_limpio}-{marca}".replace(" ", "-").lower()
    url_fallback   = f"https://listado.mercadolibre.com.ar/{slug_simple}"
    urls_extra     = []
    if url_fallback not in urls_paginadas:
        urls_extra.append(url_fallback)
    for extra in build_search_urls_extra(marca, categoria, franquicia):
        if extra not in urls_paginadas and extra not in urls_extra:
            urls_extra.append(extra)

    with sync_playwright() as p:
        session_id = f"plp_{marca}_{franquicia}_{uuid.uuid4().hex[:8]}"
        browser = p.chromium.launch(
            headless=True,
            proxy=build_proxy(session_id),
            args=["--window-size=1280,900", "--disable-web-security"],
        )
        context = setup_route_blocking(browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
            locale="es-AR",
            java_script_enabled=True,
        ))
        page = context.new_page()
        page.set_default_timeout(NAV_TIMEOUT)

        try:
            reglas_franq   = parsear_franquicia(franquicia)
            def _procesar_url(url_intento: str, idx_url: int, total_urls: int) -> int:
                """Procesa una URL de PLP y agrega resultados. Devuelve cuántos agregó."""
                nonlocal desc_franq, desc_salomon, desc_trail, desc_cross
                if MAX_PRODUCTOS_POR_FRANQUICIA > 0 and len(resultados) >= MAX_PRODUCTOS_POR_FRANQUICIA:
                    return 0

                log_scraping(f"  🔗 URL ({idx_url}/{total_urls}): {url_intento}")
                resp = page.goto(url_intento, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
                if resp and resp.status >= 400:
                    log_warning(f"  ⚠️  Status {resp.status} en intento {idx_url}")
                    return 0
                time.sleep(1.2)

                item_selector = "li.ui-search-layout__item"
                _scroll_hasta_cargar_todo(page, item_selector)
                items = page.locator(item_selector).all()
                if not items:
                    item_selector = ".ui-search-result"
                    _scroll_hasta_cargar_todo(page, item_selector)
                    items = page.locator(item_selector).all()

                log_info(f"    {len(items)} items totales tras scroll completo")
                nuevos = 0

                for item in items:
                    try:
                        link = item.locator("a").first
                        if link.count() == 0:
                            continue
                        href = link.get_attribute("href")
                        if not href or "MLA-" not in href:
                            continue

                        url_prod = href if href.startswith("http") \
                                   else f"https://www.mercadolibre.com.ar{href}"
                        mla = extract_mla_from_url(url_prod)
                        if not mla or mla in mlas_vistos:
                            continue

                        url_lower = url_prod.lower()

                        # ── Filtro: excluir Salomon y otras marcas ajenas ──────
                        if any(m in url_lower for m in MARCAS_EXCLUIR):
                            desc_salomon += 1
                            log_info(f"      ✗ Salomon/marca ajena descartado: {mla}")
                            continue

                        titulo = extract_title_from_url(url_prod)
                        try:
                            titulo_visible = item.locator(
                                "h2, .poly-component__title, .ui-search-item__title, [class*='title']"
                            ).first.inner_text().strip().lower()
                        except Exception:
                            titulo_visible = ""

                        titulo_check = (titulo_visible or titulo).lower()

                        # ── Filtro trail: excluir "trail" si franquicia no lo es ─
                        if not franq_es_trail and "trail" in titulo_check:
                            desc_trail += 1
                            log_info(f"      ✗ Trail descartado: {mla}")
                            continue

                        # ── Filtro FUTBOL: excluir zapatillas ─────────────────
                        if cat_upper == "FUTBOL":
                            if "zapatilla" in titulo_check:
                                desc_cross += 1
                                log_info(f"      ✗ Zapatilla descartada en FUTBOL: {mla}")
                                continue

                        # ── Filtro Running/Training/Sportswear: excluir botines/fútbol ─
                        if cat_lower in ("running", "training", "sportswear"):
                            if any(w in titulo_check for w in ("botin", "fútbol", "futbol", "soccer")):
                                desc_cross += 1
                                log_info(f"      ✗ Botín/fútbol descartado en {categoria}: {mla}")
                                continue

                        # ── Filtro accesorios: excluir no-calzado en categorías calzado ─
                        if cat_lower in ("running", "training", "sportswear", "futbol"):
                            if any(w in titulo_check for w in PALABRAS_ACCESORIOS):
                                desc_cross += 1
                                log_info(f"      ✗ Accesorio descartado en {categoria}: {mla}")
                                continue

                        # ── Filtro: franquicia en título ───────────────────────
                        if not match_franquicia(titulo, reglas_franq):
                            if titulo_visible and match_franquicia(titulo_visible, reglas_franq):
                                log_info(f"      ↩ Fallback título visible: '{titulo_visible[:50]}'")
                                titulo = titulo_visible
                            else:
                                desc_franq += 1
                                continue

                        mlas_vistos.add(mla)
                        resultados.append({
                            "mla": mla,
                            "url": url_prod,
                            "marca": marca,
                            "categoria": categoria,
                            "franquicia": franquicia,
                            "product_name_plp": titulo_visible or titulo,
                        })
                        nuevos += 1
                        if MAX_PRODUCTOS_POR_FRANQUICIA > 0 and len(resultados) >= MAX_PRODUCTOS_POR_FRANQUICIA:
                            break

                    except Exception:
                        continue

                return nuevos

            # ── Páginas 1..MAX_PAGINAS de la query principal ──────────────────
            total_urls_main = len(urls_paginadas)
            for idx, url in enumerate(urls_paginadas, start=1):
                if MAX_PRODUCTOS_POR_FRANQUICIA > 0 and len(resultados) >= MAX_PRODUCTOS_POR_FRANQUICIA:
                    break
                antes = len(resultados)
                _procesar_url(url, idx, total_urls_main + len(urls_extra))
                nuevos_en_pag = len(resultados) - antes
                # Si la pág 1 trajo 0 resultados, no tiene sentido paginar más
                if idx == 1 and nuevos_en_pag == 0:
                    log_warning(f"  ⚠️  Pág 1 sin resultados — saltando paginación")
                    break
                # Si la página siguiente no trae nada nuevo, parar paginación
                if idx > 1 and nuevos_en_pag == 0:
                    log_info(f"  Página {idx} sin nuevos items — fin paginación")
                    break

            # ── Fallback URLs (Club, Revolution, zoom, air, etc.) ─────────────
            offset_extra = total_urls_main
            for idx, url in enumerate(urls_extra, start=offset_extra + 1):
                if MAX_PRODUCTOS_POR_FRANQUICIA > 0 and len(resultados) >= MAX_PRODUCTOS_POR_FRANQUICIA:
                    break
                _procesar_url(url, idx, total_urls_main + len(urls_extra))

        except Exception as e:
            log_error(f"  Error en PLP tienda oficial: {e}")
        finally:
            browser.close()

    log_success(
        f"  ✅ {len(resultados)} productos — descartados: "
        f"{desc_franq} franquicia, {desc_salomon} marcas ajenas, "
        f"{desc_trail} trail, {desc_cross} cross-cat/accesorios"
    )
    return resultados


def _resolver_overlap_franquicias(productos: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Cuando un mismo MLA aparece en dos franquicias (ej: PEGASUS y PEGASUS PLUS),
    lo asigna SOLO a la franquicia más específica (nombre más largo / más palabras).
    Ej: MLA-1234 con slug 'pegasus plus' → va a PEGASUS PLUS, se elimina de PEGASUS.
    """
    from collections import defaultdict

    # Agrupar por MLA
    mla_franqs: dict = defaultdict(list)
    for prod in productos:
        mla_franqs[prod["mla"]].append(prod)

    resultado = []
    for mla, prods in mla_franqs.items():
        if len(prods) == 1:
            resultado.append(prods[0])
            continue

        # MLA duplicado — elegir la franquicia más específica
        # Extraer slug del título desde la URL
        url = prods[0].get("url", "")
        titulo = extract_title_from_url(url)

        # Ordenar franquicias de más palabras a menos (más específica primero)
        # Ordenar por especificidad: más palabras obligatorias primero,
        # luego más palabras en total
        def _especificidad(p):
            r = parsear_franquicia(p["franquicia"])
            return (len(r["palabras_obligatorias"]) + (1 if r["frase_exacta"] else 0),
                    len(r["palabras_libres"]))
        prods_sorted = sorted(prods, key=_especificidad, reverse=True)

        # Elegir la primera que matchee el título
        elegida = None
        for prod in prods_sorted:
            reglas = parsear_franquicia(prod["franquicia"])
            if match_franquicia(titulo, reglas):
                elegida = prod
                break

        if not elegida:
            elegida = prods_sorted[0]  # fallback: la más larga

        descartadas = [p["franquicia"] for p in prods if p is not elegida]
        log_info(f"  🔀 MLA {mla}: asignado a '{elegida['franquicia']}' "
                 f"(descartado de: {descartadas})")
        resultado.append(elegida)

    return resultado


def collect_all_plps_threaded(
    franquicias: List[Dict[str, str]],
    max_workers: int = PLP_WORKERS,
) -> List[Dict[str, str]]:
    """Scrapea todas las PLPs en paralelo con ThreadPoolExecutor."""
    log_info(f"\n{'='*60}")
    log_info(f"🔍 SCRAPEO PLPs — {max_workers} threads — {len(franquicias)} búsquedas")
    log_info(f"{'='*60}")

    todos: List[Dict[str, str]] = []
    lock  = threading.Lock()
    start = datetime.now()

    def _worker(f: Dict[str, str]):
        prods = scrape_plp_for_franchise(f["marca"], f["categoria"], f["franquicia"])
        with lock:
            todos.extend(prods)
            elapsed = (datetime.now() - start).total_seconds()
            log_info(f"  Acumulado: {len(todos)} productos ({elapsed:.0f}s)")

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(_worker, f) for f in franquicias]
        for fut in as_completed(futures):
            try:
                fut.result()
            except Exception as e:
                log_error(f"Error en thread PLP: {e}")

    elapsed = (datetime.now() - start).total_seconds()
    log_success(f"\n✅ TOTAL PLP: {len(todos)} productos en {elapsed:.0f}s")
    todos = _resolver_overlap_franquicias(todos)
    log_success(f"✅ Post-dedup overlaps: {len(todos)} productos únicos")
    return todos

# ============================================================
# SCRAPEO DE PDP
# ============================================================

def scrape_pdp(producto: Dict[str, str]) -> Optional[Dict[str, Any]]:
    """
    Scrapea la página de detalle (PDP) de un producto.
    Extrae: full_price, final_price, cuotas, talles.
    """
    with sync_playwright() as p:
        session_id = f"pdp_{producto['mla']}_{uuid.uuid4().hex[:8]}"
        browser = p.chromium.launch(
            headless=True,
            proxy=build_proxy(session_id),
            args=["--window-size=1280,900", "--disable-web-security"],
        )
        context = setup_route_blocking(browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
            locale="es-AR",
            java_script_enabled=True,
        ))
        page = context.new_page()
        page.set_default_timeout(NAV_TIMEOUT)

        try:
            log_scraping(f"    PDP: {producto['mla']}")
            page.goto(producto["url"], wait_until="domcontentloaded")

            # Esperar que los componentes React de variantes estén montados.
            # Meli carga el layout primero y monta talles/colores de forma lazy.
            # Señal de que está listo: el precio final aparece en el DOM.
            try:
                page.wait_for_selector(
                    'meta[itemprop="price"], .andes-money-amount__fraction',
                    timeout=10000
                )
            except Exception:
                pass
            time.sleep(1.5)  # buffer extra para el montaje de variantes

            # ── Filtro 1: solo productos NUEVOS ──────────────────────────
            try:
                condicion_el = page.locator(".ui-pdp-subtitle").first
                if condicion_el.count() > 0:
                    condicion = condicion_el.inner_text().lower()
                    if "usado" in condicion or "reacondicionado" in condicion:
                        log_warning(f"    ✗  Producto usado/reacondicionado — saltando")
                        return None
            except Exception:
                pass

            # ── Filtro 2: verificar vendedor es tienda oficial ────────────
            # Solo descarta si CONFIRMA que el vendedor NO es Nike/Adidas/Puma.
            # Si no puede leer el vendedor → acepta (no descartar por las dudas).
            try:
                nombres_oficiales = {
                    "nike":   ["nike"],
                    "adidas": ["adidas"],
                    "puma":   ["puma"],
                }
                marca_lower = producto["marca"].lower()
                nombres_ok  = nombres_oficiales.get(marca_lower, [])

                # Selectores del bloque de vendedor en Meli PDP
                vendedor_selectors = [
                    ".ui-pdp-seller__header__title",
                    ".ui-pdp-seller__header .ui-pdp-color--BLUE",
                    "[class*='seller__header'] a",
                    "[class*='seller__link']",
                    ".ui-pdp-action-modal__title",
                ]
                vendedor_confirmado = False
                for sel in vendedor_selectors:
                    try:
                        el = page.locator(sel).first
                        if el.count() == 0:
                            continue
                        vtext = (el.inner_text() or "").lower().strip()
                        if not vtext:
                            continue
                        # ¿Es la tienda oficial de la marca?
                        if any(n in vtext for n in nombres_ok):
                            log_info(f"    ✅ Tienda oficial: '{vtext}'")
                            vendedor_confirmado = True
                            break
                        # ¿Es claramente otro vendedor (texto largo = nombre de tienda)?
                        if len(vtext) > 3 and not any(n in vtext for n in nombres_ok):
                            log_warning(f"    ✗  Vendedor '{vtext}' no es tienda oficial — saltando")
                            return None
                    except Exception:
                        continue
                # Si no pudo leer el vendedor → no descartar
                if not vendedor_confirmado:
                    log_info(f"    ⚠️  Vendedor no determinado — aceptando")
            except Exception:
                pass

            full_price, final_price, full_raw, final_raw = extract_prices(page)
            if not full_price or not final_price:
                log_warning(f"    Sin precios — saltando {producto['mla']}")
                return None

            # Sanity check: si final > full es error del scraper (precio tachado mal leído)
            # → tomamos full = final para que markdown quede 0 en lugar de negativo
            if final_price > full_price:
                log_warning(f"    ⚠️  final (${final_price:,.0f}) > full (${full_price:,.0f}) "
                            f"— ajustando full = final, markdown = 0%")
                full_price = final_price
                full_raw   = final_raw

            cuotas = extract_installments(page, final_price)
            talles = extract_sizes(page)

            resultado = {
                "mla":          producto["mla"],
                "url":          producto["url"],
                "marca":        producto["marca"],
                "categoria":    producto["categoria"],
                "franquicia":   producto["franquicia"],
                "product_name": (page.locator("h1.ui-pdp-title, h1").first.inner_text().strip()
                                  if page.locator("h1.ui-pdp-title, h1").first.count() > 0
                                  else str(producto.get("product_name_plp", "")).strip()),
                "full_price":   full_price,
                "final_price":  final_price,
                "full_raw":     full_raw,
                "final_raw":    final_raw,
                "cuotas":       cuotas,
                "talles":       talles,
                "markdown_pct": (full_price - final_price) / full_price
                                if full_price > 0 else 0,
                "fecha":        datetime.now().strftime("%Y-%m-%d"),
            }
            log_success(f"    ✅ {producto['mla']} — ${final_price:,.0f} — {talles} talles")
            return resultado

        except Exception as e:
            log_error(f"    ❌ Error {producto['mla']}: {e}")
            return None
        finally:
            browser.close()


def worker_process_pdps(chunk: List[Dict[str, str]], worker_id: int) -> List[Dict[str, Any]]:
    """Worker para un lote de PDPs (corre en proceso separado)."""
    resultados: List[Dict[str, Any]] = []
    log_info(f"  Worker {worker_id}: {len(chunk)} PDPs")
    for i, prod in enumerate(chunk, 1):
        try:
            log_info(f"    Worker {worker_id} — {i}/{len(chunk)}")
            r = scrape_pdp(prod)
            if r:
                resultados.append(r)
            time.sleep(0.5)
        except Exception as e:
            log_error(f"  Worker {worker_id} error: {e}")
    log_info(f"  Worker {worker_id} terminó: {len(resultados)} OK")
    return resultados


def scrape_all_pdps_parallel(
    productos: List[Dict[str, str]],
    max_workers: int = PDP_WORKERS,
) -> List[Dict[str, Any]]:
    """Scrapea todas las PDPs divididas en chunks por proceso."""
    log_info(f"\n{'='*60}")
    log_info(f"📦 SCRAPEO PDPs — {max_workers} procesos — {len(productos)} productos")
    log_info(f"{'='*60}")
    if not productos:
        return []

    start      = datetime.now()
    chunk_size = max(1, len(productos) // (max_workers * 2))
    chunks     = [productos[i:i + chunk_size]
                  for i in range(0, len(productos), chunk_size)]
    log_info(f"  {len(chunks)} chunks de ~{chunk_size} productos")

    todos: List[Dict[str, Any]] = []
    with ProcessPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(worker_process_pdps, c, i + 1): i
                   for i, c in enumerate(chunks)}
        for fut in as_completed(futures):
            try:
                res = fut.result(timeout=300)
                todos.extend(res)
                elapsed = (datetime.now() - start).total_seconds()
                log_info(f"  Progreso: {len(todos)} scrapeados ({elapsed:.0f}s)")
            except Exception as e:
                log_error(f"Error en worker PDP: {e}")

    elapsed = (datetime.now() - start).total_seconds()
    log_success(f"\n✅ TOTAL PDPs: {len(todos)} en {elapsed:.0f}s")
    return todos

def rescrape_talles_sospechosos(
    resultados: List[Dict[str, Any]],
    max_workers: int = PDP_WORKERS,
) -> List[Dict[str, Any]]:
    """
    Segunda pasada de PDP solo para productos que quedaron con talles == 1.

    Cuando un producto tiene exactamente 1 talle scrapeado, es muy probable que
    el dropdown no se haya abierto correctamente en la primera pasada (problema
    intermitente con dropdowns de Meli). Se re-scrapea el PDP completo y, si
    el nuevo resultado trae más talles, se reemplaza en la lista.

    Productos con talles == 0 NO se re-scrapean (pueden ser indumentaria sin talles).
    """
    sospechosos = [r for r in resultados if r.get("talles", 0) == 1]
    if not sospechosos:
        log_info("  ✅ No hay productos con 1 solo talle — no se necesita re-scraping")
        return resultados

    log_info(f"\n{'='*60}")
    log_info(f"🔁 RE-SCRAPING TALLES — {len(sospechosos)} productos con talles==1")
    log_info(f"{'='*60}")
    for s in sospechosos:
        log_info(f"  → {s['mla']} ({s['marca']} / {s['franquicia']})")

    # Convertir lista de resultados a dict para fácil reemplazo por MLA
    resultados_dict = {r["mla"]: r for r in resultados}

    # Re-scrapear en paralelo (misma lógica que el pipeline principal)
    productos_input = [
        {"mla": r["mla"], "url": r["url"],
         "marca": r["marca"], "categoria": r["categoria"],
         "franquicia": r["franquicia"]}
        for r in sospechosos
    ]

    chunk_size = max(1, len(productos_input) // (max_workers * 2))
    chunks = [productos_input[i:i + chunk_size]
              for i in range(0, len(productos_input), chunk_size)]

    mejorados = 0
    with ProcessPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(worker_process_pdps, c, i + 1): i
                   for i, c in enumerate(chunks)}
        for fut in as_completed(futures):
            try:
                nuevos = fut.result(timeout=300)
                for nuevo in nuevos:
                    mla = nuevo["mla"]
                    talles_anterior = resultados_dict[mla]["talles"]
                    talles_nuevo    = nuevo["talles"]
                    if talles_nuevo > talles_anterior:
                        log_success(
                            f"  🔁 {mla}: talles {talles_anterior} → {talles_nuevo} ✅"
                        )
                        resultados_dict[mla] = nuevo
                        mejorados += 1
                    else:
                        log_info(
                            f"  🔁 {mla}: sin mejora ({talles_anterior} → {talles_nuevo})"
                        )
            except Exception as e:
                log_error(f"  Error en re-scraping talles: {e}")

    log_success(f"  ✅ Re-scraping talles: {mejorados}/{len(sospechosos)} productos mejorados")
    return list(resultados_dict.values())




def generar_output(resultados: List[Dict[str, Any]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Devuelve dos DataFrames:
      1. Agrupado por (categoria, marca, franquicia) con promedios
      2. Crudo con todos los productos individuales
    Ambos ordenados por categoría y precio final ascendente.
    """
    grupos = defaultdict(lambda: {
        "precios_full": [], "precios_final": [],
        "cuotas": [], "talles": [], "urls": [],
    })
    for r in resultados:
        key = (r["categoria"], r["marca"], r["franquicia"])
        grupos[key]["precios_full"].append(r["full_price"])
        grupos[key]["precios_final"].append(r["final_price"])
        grupos[key]["cuotas"].append(r["cuotas"])
        grupos[key]["talles"].append(r["talles"])
        grupos[key]["urls"].append(r["url"])

    rows_agrupado = []
    for (cat, marca, franq), data in grupos.items():
        pf = data["precios_final"]
        if not pf:
            continue
        full_prom     = sum(data["precios_full"]) / len(data["precios_full"])
        final_prom    = sum(pf) / len(pf)
        markdown_prom = (full_prom - final_prom) / full_prom if full_prom > 0 else 0
        # Talles: excluir None (dropdown ilegible) del promedio
        talles_validos = [t for t in data["talles"] if t is not None]
        talles_prom = round(sum(talles_validos) / len(talles_validos), 1) if talles_validos else ""

        # Cuotas: excluir None y 0 del promedio
        cuotas_validas = [c for c in data["cuotas"] if c]
        cuotas_prom = round(sum(cuotas_validas) / len(cuotas_validas), 1) if cuotas_validas else ""

        rows_agrupado.append({
            "Categoría":           cat,
            "Marca":               marca.capitalize(),
            "Franquicia":          franq,
            "Cantidad MLA":        len(pf),
            "Full Price Promedio": f"$ {full_prom:,.0f}".replace(",", "."),
            "Final Price Promedio":f"$ {final_prom:,.0f}".replace(",", "."),
            "Markdown Promedio %": f"{markdown_prom * 100:.1f}%",
            "Cuotas Promedio":     cuotas_prom,
            "Talles Promedio x MLA": talles_prom,
            "URL Ejemplo":         data["urls"][0] if data["urls"] else "",
        })

    df_agrupado = pd.DataFrame(rows_agrupado)
    if not df_agrupado.empty:
        df_agrupado["_precio_num"] = (
            df_agrupado["Final Price Promedio"]
            .str.replace("$ ", "", regex=False)
            .str.replace(".", "", regex=False)
            .astype(float)
        )
        df_agrupado = (df_agrupado
                       .sort_values(["Categoría", "_precio_num"])
                       .drop(columns=["_precio_num"])
                       .reset_index(drop=True))

    rows_crudo = []
    for r in resultados:
        fp = r["full_price"]
        fn = r["final_price"]
        rows_crudo.append({
            "Categoría":  r["categoria"],
            "Marca":      r["marca"].capitalize(),
            "Franquicia": r["franquicia"],
            "MLA":        r["mla"],
            "Full Price": f"$ {fp:,.0f}".replace(",", "."),
            "Final Price":f"$ {fn:,.0f}".replace(",", "."),
            "Markdown %": f"{(fp - fn) / fp * 100:.1f}%" if fp > 0 else "",
            "Cuotas":     r["cuotas"],
            "Talles":     r["talles"] if r["talles"] is not None else "",
            "URL":        r["url"],
            "Fecha":      r["fecha"],
            "Nombre Producto": r.get("product_name", ""),
        })

    df_crudo = pd.DataFrame(rows_crudo)
    return df_agrupado, df_crudo

# ============================================================
# EXCEL WRITER
# ============================================================

def write_excel(df_agrupado: pd.DataFrame, df_crudo: pd.DataFrame, output_path: str):
    """Escribe Excel con dos solapas y formato profesional."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_agrupado.to_excel(writer, sheet_name="Resumen por Franquicia", index=False)
        df_crudo.to_excel(writer,    sheet_name="Todos los Productos",    index=False)

        wb          = writer.book
        hdr_font    = Font(bold=True, color="FFFFFF")
        hdr_fill    = PatternFill("solid", fgColor="1E4C7A")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        for sheet_name in ("Resumen por Franquicia", "Todos los Productos"):
            ws = wb[sheet_name]
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.border    = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Autofit ancho de columnas (máximo 50)
            for col in range(1, ws.max_column + 1):
                max_len = max(
                    len(str(ws.cell(row=r, column=col).value or ""))
                    for r in range(1, min(ws.max_row + 1, 100))
                )
                ws.column_dimensions[get_column_letter(col)].width = min(50, max_len + 2)

# ============================================================
# CACHE
# ============================================================

class SimpleCache:
    """Cache JSON por marca, sin TTL."""

    def __init__(self, cache_dir: str = CACHE_DIR):
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(exist_ok=True)
        self.cache: Dict[str, Dict] = {}
        self._load_all()

    def _path(self, marca: str) -> Path:
        return self.cache_dir / f"cache_{marca.lower()}.json"

    def _load_all(self):
        for f in self.cache_dir.glob("cache_*.json"):
            marca = f.stem.replace("cache_", "")
            try:
                self.cache[marca] = json.loads(f.read_text(encoding="utf-8"))
            except Exception:
                self.cache[marca] = {}
        log_info(f"📦 Cache cargado: {sum(len(v) for v in self.cache.values())} entradas")

    def get(self, marca: str, mla: str) -> Optional[Dict]:
        return self.cache.get(marca, {}).get(mla)

    def set(self, marca: str, mla: str, data: Dict):
        self.cache.setdefault(marca, {})[mla] = data

    def save_all(self):
        for marca, data in self.cache.items():
            self._path(marca).write_text(
                json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
            )
        log_info("💾 Cache guardado")

# ============================================================
# MAIN
# ============================================================

def main():
    # ── Filtro de prueba ─────────────────────────────────────────────────────
    # Cambiar CATEGORIA_FILTRO a "Running" (o cualquier categoría) para prueba parcial
    # None = corre TODAS las categorías (modo producción)
    CATEGORIA_FILTRO     = None
    MAX_PRODUCTOS_PRUEBA = 100
    # ────────────────────────────────────────────────────────────────────────

    modo_prueba = CATEGORIA_FILTRO is not None

    print("\n" + "=" * 90)
    print("🚀  MELI MULTI-BRAND SCRAPER — NIKE VS ADIDAS VS PUMA  v11")
    if modo_prueba:
        print(f"  🧪  MODO PRUEBA — categoría: {CATEGORIA_FILTRO} "
              f"(max {MAX_PRODUCTOS_PRUEBA} productos/franquicia)")
    else:
        print("  🏭  MODO PRODUCCIÓN — todas las categorías")
    print("=" * 90)
    print(f"  📊  Excel     : {DEFAULT_EXCEL_PATH}")
    print(f"  📡  Proxy     : {PROXY_HOST}:{PROXY_PORT}")
    print(f"  👷  PLP workers: {PLP_WORKERS} threads")
    print(f"  👷  PDP workers: {PDP_WORKERS} procesos")
    print(f"  📦  Cache     : {CACHE_DIR}")
    print("=" * 90 + "\n")

    t0 = datetime.now()

    # Test proxy
    log_info("🔌 Testeando proxy...")
    proxy_ok, working_port = test_proxy_simple()
    if not proxy_ok:
        resp = input("\n¿Continuar sin proxy? (s/n): ")
        if resp.strip().lower() != "s":
            return
    else:
        log_success(f"Proxy OK en puerto {working_port}")

    # Diagnóstico del Excel
    log_info("🔎 Diagnóstico del Excel...")
    try:
        from openpyxl import load_workbook as _lw
        _wb = _lw(DEFAULT_EXCEL_PATH, data_only=True)
        log_info(f"  Solapas: {_wb.sheetnames}")
        _ws = _wb.active
        log_info(f"  Solapa activa: '{_ws.title}' — {_ws.max_row} filas x {_ws.max_column} cols")
        log_info("  Primeras 5 filas (cols 1-5):")
        for r in range(1, min(6, _ws.max_row + 1)):
            fila = [repr(_ws.cell(row=r, column=c).value) for c in range(1, min(6, _ws.max_column + 1))]
            log_info(f"    fila {r}: {' | '.join(fila)}")
    except Exception as e:
        log_error(f"  Error en diagnóstico Excel: {e}")

    # Leer franquicias
    try:
        franquicias = read_franchises_from_excel(DEFAULT_EXCEL_PATH)
    except Exception as e:
        log_error(f"Error leyendo Excel: {e}")
        return
    if not franquicias:
        log_error("No se encontraron franquicias en el Excel.")
        return

    if not SCRAP_PUMA:
        antes = len(franquicias)
        franquicias = [f for f in franquicias if f.get("marca", "").strip().lower() != "puma"]
        excluidas = antes - len(franquicias)
        log_info(f"🐆 SCRAP_PUMA=False → {excluidas} franquicias Puma excluidas")

    # Filtro de categoría
    if CATEGORIA_FILTRO:
        franquicias = [f for f in franquicias
                       if f["categoria"].strip().lower() == CATEGORIA_FILTRO.strip().lower()]
        if not franquicias:
            log_error(f"No hay franquicias para: {CATEGORIA_FILTRO}")
            return
        log_success(f"Filtro: {len(franquicias)} franquicias de '{CATEGORIA_FILTRO}'")
        for f in franquicias:
            log_info(f"  → {f['marca'].capitalize()} / {f['franquicia']}")

    # Reducir límite en modo prueba
    global MAX_PRODUCTOS_POR_FRANQUICIA
    if modo_prueba:
        MAX_PRODUCTOS_POR_FRANQUICIA = MAX_PRODUCTOS_PRUEBA
        log_info(f"  Límite: {MAX_PRODUCTOS_PRUEBA} productos/franquicia")

    cache = SimpleCache()

    # PLP
    todos_productos = collect_all_plps_threaded(franquicias)
    if not todos_productos:
        log_error("No se encontraron productos en PLPs.")
        return

    # PDP
    resultados = scrape_all_pdps_parallel(todos_productos)
    if not resultados:
        log_error("No se pudieron scrapear PDPs.")
        return

    # talles == 0   → sin picker → sacar del output
    # talles == None → dropdown ilegible → mantener con campo vacío (excluido del promedio)
    # talles >= 1   → real → mantener
    antes = len(resultados)
    resultados = [r for r in resultados if r.get("talles") != 0]
    eliminados = antes - len(resultados)
    if eliminados > 0:
        log_warning(f"  🗑️  {eliminados} productos sin picker de talles eliminados del output")

    # Guardar en cache
    for prod in resultados:
        cache.set(prod["marca"], prod["mla"], prod)
    cache.save_all()

    # Generar output
    df_agrupado, df_crudo = generar_output(resultados)

    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path   = f"meli_nike_adidas_puma_{ts}.csv"
    excel_path = f"meli_nike_adidas_puma_{ts}.xlsx"

    df_crudo.to_csv(csv_path, index=False, encoding="utf-8-sig")
    write_excel(df_agrupado, df_crudo, excel_path)

    elapsed = (datetime.now() - t0).total_seconds()
    print("\n" + "=" * 90)
    log_success("PIPELINE COMPLETADO")
    print("=" * 90)
    log_info(f"⏱️  Tiempo total      : {elapsed:.1f}s")
    log_info(f"📋  Franquicias       : {len(franquicias)}")
    log_info(f"🔍  URLs de PLP       : {len(todos_productos)}")
    log_info(f"📦  PDPs exitosas     : {len(resultados)}")
    log_info(f"📄  CSV               : {csv_path}")
    log_info(f"📄  Excel             : {excel_path}")
    print("=" * 90 + "\n")

    if not df_agrupado.empty:
        print("\n📋 TOP 10 FRANQUICIAS (por precio):")
        cols = ["Categoría", "Marca", "Franquicia", "Cantidad MLA",
                "Final Price Promedio", "Cuotas Promedio", "Talles Promedio x MLA"]
        print(df_agrupado[cols].head(10).to_string(index=False))


if __name__ == "__main__":
    main()