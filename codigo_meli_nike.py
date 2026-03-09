# -*- coding: utf-8 -*-
"""
Mercado Libre (MeLi) scraper for Nike Southbay – Versión Final con Reintentos y Cache de 4 días
"""

import os
import time
import random

# --- Hard limit BLAS/NumExpr threads ---
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("VECLIB_MAXIMUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_MAX_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")

import uuid
import re
import sys
import json
import math
import argparse
import datetime as dt
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from concurrent.futures import ProcessPoolExecutor, as_completed
import base64
import socket
import ssl
import urllib3
from urllib3.exceptions import MaxRetryError, ProxyError as Urllib3ProxyError
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import pandas as pd
from pyxlsb import open_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError, Error as PWError

# Deshabilitar warnings de SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import warnings
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

DEFAULT_SEASON = "SP26"
COMPETITOR_NAME = "Mercado Libre"

# Shipping (ARS)
MELI_STD_SHIPPING_ARS = 6000
NIKE_STD_SHIPPING_ARS = 10899
MELI_FREE_SHIP_FROM_ARS = 35000
NIKE_FREE_SHIP_FROM_ARS = 99999

# Nike cuotas model
NIKE_CUOTAS_SIMPLE_MODE = True
NIKE_CUOTAS_ALL = 3
NIKE_CUOTAS_HIGH = 6
NIKE_CUOTAS_HIGH_FROM_ARS = 20000

# Archivos y paths
DEFAULT_STATUSBOOK_PATH = r"StatusBooks NDDC ARG SP26.xlsb"
STATUSBOOK_SHEET = "Books NDDC"
STATUSBOOK_HEADER_ROW_IDX = 6  # 0-index (fila 7)

# Tolerancias
BML_THRESHOLD_PCT = 0.02        # ±2% para Beat/Meet/Lose
CUOTAS_TOL_PCT = 0.02           # ±2% para validar cuotas sin interés

# Cache - 4 días de TTL
CACHE_DEFAULT_PATH = "meli_cache.json"
CACHE_TTL_DAYS = 0
REFRESH_CACHED_DEFAULT = False

# Playwright
DEFAULT_HEADLESS = False
DEFAULT_NAV_TIMEOUT_MS = 90_000

# Rotación de IP por bloqueo
ROTATE_IP_ON_WALL = True
MAX_IP_ROTATIONS_PER_ITEM = 3
ROTATION_BACKOFF_MS = 1200

# Workers - 5 workers fijos
DEFAULT_WORKERS = 2
MAX_SAFE_WORKERS = 3

# Reintentos para errores SSL y navegación
MAX_SCRAPE_RETRIES = 3
RETRY_BACKOFF_FACTOR = 2  # 2, 4, 8 segundos
RETRY_ON_ERRORS = [
    "ERR_TUNNEL_CONNECTION_FAILED",
    "ERR_PROXY_CONNECTION_FAILED",
    "ERR_SSL_PROTOCOL_ERROR",
    "ERR_SSL_VERSION_OR_CIPHER_MISMATCH",
    "ERR_CONNECTION_REFUSED",
    "ERR_CONNECTION_RESET",
    "ERR_CONNECTION_CLOSED",
    "ERR_NAME_NOT_RESOLVED",
    "ERR_INTERNET_DISCONNECTED",
    "navigation interrupted",
    "net::ERR_",
    "Timeout"
]

REBUILD_ON_TUNNEL_ERRORS = True  # Rebuild browser/context ante ERR_TUNNEL / ERR_PROXY / SSL
TUNNEL_ERROR_PATTERNS = [
    "ERR_TUNNEL_CONNECTION_FAILED",
    "ERR_PROXY_CONNECTION_FAILED",
    "ERR_SSL_PROTOCOL_ERROR",
]

# ============================================================
# DECODO PROXY CONFIGURATION
# ============================================================

# ============================================================
# DECODO PROXY CONFIGURATION (MOBILE)
# ============================================================

# Decodo Mobile proxies usan gate.decodo.com + puertos por endpoint
DECODO_HOST = "gate.decodo.com"

# En tu panel tenés 10 endpoints: 10001..10010
DECODO_PORTS = list(range(10001, 10011))

# Credenciales Mobile (del panel)
DECODO_USER = "spyrndvq0x"
DECODO_PASS = "8eOzLZZj3i3b=mcoc8"

# Mantengo USERNAME_VARIANTS para NO tocar la lógica de test híbrido
# (pero en Mobile no hace falta probar variantes: es 1 solo username)
USERNAME_VARIANTS = [DECODO_USER]

# ============================================================
# LOGGING ULTRA DETALLADO
# ============================================================

DEBUG_MODE = True
COLORS = {
    "HEADER": "\033[95m",
    "BLUE": "\033[94m",
    "GREEN": "\033[92m",
    "YELLOW": "\033[93m",
    "RED": "\033[91m",
    "CYAN": "\033[96m",
    "WHITE": "\033[97m",
    "BOLD": "\033[1m",
    "UNDERLINE": "\033[4m",
    "END": "\033[0m"
}

def log_info(msg: str):
    timestamp = dt.datetime.now().strftime("%H:%M:%S")
    print(f"{COLORS['GREEN']}[{timestamp} INFO]{COLORS['END']} {msg}")

def log_warning(msg: str):
    timestamp = dt.datetime.now().strftime("%H:%M:%S")
    print(f"{COLORS['YELLOW']}[{timestamp} WARN]{COLORS['END']} {msg}")

def log_error(msg: str):
    timestamp = dt.datetime.now().strftime("%H:%M:%S")
    print(f"{COLORS['RED']}[{timestamp} ERROR]{COLORS['END']} {msg}")

def log_success(msg: str):
    timestamp = dt.datetime.now().strftime("%H:%M:%S")
    print(f"{COLORS['GREEN']}[{timestamp} OK]{COLORS['END']} {msg}")

def log_debug(msg: str):
    if DEBUG_MODE:
        timestamp = dt.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        print(f"{COLORS['CYAN']}[{timestamp} DEBUG]{COLORS['END']} {msg}")

def log_scraping(msg: str):
    timestamp = dt.datetime.now().strftime("%H:%M:%S")
    print(f"{COLORS['BLUE']}[{timestamp} SCRAPE]{COLORS['END']} {msg}")

def log_proxy(msg: str):
    timestamp = dt.datetime.now().strftime("%H:%M:%S")
    print(f"{COLORS['YELLOW']}[{timestamp} PROXY]{COLORS['END']} {msg}")

def log_http(msg: str, status_code: int = None):
    timestamp = dt.datetime.now().strftime("%H:%M:%S")
    if status_code:
        if status_code >= 500:
            color = COLORS['RED']
        elif status_code >= 400:
            color = COLORS['YELLOW']
        else:
            color = COLORS['GREEN']
        print(f"{color}[{timestamp} HTTP {status_code}]{COLORS['END']} {msg}")
    else:
        print(f"{COLORS['CYAN']}[{timestamp} HTTP]{COLORS['END']} {msg}")

# ============================================================
# FUNCIONES DE PROXY HÍBRIDAS
# ============================================================


def decodo_build_proxy(
    session_id: Optional[str] = None,
    country: str = "ar",
    username_variant: Optional[str] = None,
    forced_port: Optional[int] = None
) -> Dict[str, str]:
    """
    Construye configuración de proxy para Decodo Mobile.
    - Server: http://gate.decodo.com:<port>
    - Username/Password: los del panel Mobile
    - En Mobile la "sticky session" suele estar atada al puerto/endpoint.
      Por eso soportamos:
        * forced_port (prioridad absoluta): fija el endpoint del worker.
        * session_id: fallback para mapear a un puerto de forma estable.
    """
    # En Mobile no usamos country routing via username. Dejamos el parámetro para no romper firmas.
    username = (username_variant or DECODO_USER)

    # 1) forced_port tiene prioridad (evita colisiones entre workers)
    if forced_port is not None:
        port = int(forced_port)
        if port not in DECODO_PORTS:
            # si viene un puerto fuera del rango, lo normalizamos a uno válido
            port = DECODO_PORTS[int(abs(port)) % len(DECODO_PORTS)]
    else:
        # 2) session_id => mapeo estable a un puerto (para sticky)
        if session_id:
            try:
                idx = abs(hash(session_id)) % len(DECODO_PORTS)
            except Exception:
                idx = 0
            port = DECODO_PORTS[idx]
        else:
            # 3) default: primer puerto
            port = DECODO_PORTS[0]

    server = f"http://{DECODO_HOST}:{port}"

    return {
        "server": server,
        "username": username,
        "password": DECODO_PASS
    }
def test_proxy_playwright(country: str = "ar", username_variant: Optional[str] = None) -> Tuple[bool, Optional[Dict], str]:
    """Prueba el proxy usando Playwright."""
    try:
        from playwright.sync_api import sync_playwright
        
        with sync_playwright() as p:
            proxy = decodo_build_proxy(country=country, username_variant=username_variant)
            log_debug(f"  Probando Playwright con username: {proxy['username']}")
            
            for browser_type in ['chromium', 'webkit', 'firefox']:
                try:
                    log_debug(f"    Navegador: {browser_type}")
                    
                    if browser_type == 'chromium':
                        browser = p.chromium.launch(
                            headless=True,
                            proxy=proxy,
                            args=["--ignore-certificate-errors"]
                        )
                    elif browser_type == 'webkit':
                        browser = p.webkit.launch(headless=True, proxy=proxy)
                    else:
                        browser = p.firefox.launch(headless=True, proxy=proxy)
                    
                    page = browser.new_page()
                    response = page.goto("https://ip.decodo.com/json", timeout=15000)
                    
                    if response and response.status == 200:
                        content = page.inner_text("body")
                        data = json.loads(content)
                        browser.close()
                        log_success(f"    ✅ Playwright ({browser_type}) OK")
                        return True, data, f"playwright_{browser_type}"
                    
                    browser.close()
                    
                except Exception as e:
                    log_debug(f"    {browser_type} falló: {str(e)[:100]}")
                    continue
                    
    except Exception as e:
        log_debug(f"  Playwright general falló: {e}")
    
    return False, None, "playwright_none"

def test_proxy_urllib3(country: str = "ar", username_variant: Optional[str] = None) -> Tuple[bool, Optional[Dict], str]:
    """Prueba el proxy usando urllib3."""
    try:
        proxy = decodo_build_proxy(country=country, username_variant=username_variant)
        proxy_url = f"http://{proxy['username']}:{proxy['password']}@{proxy['server'].replace('http://', '')}"
        
        log_debug(f"  Probando urllib3 con username: {proxy['username']}")
        
        auth = base64.b64encode(f"{proxy['username']}:{proxy['password']}".encode()).decode()
        
        http = urllib3.ProxyManager(
            proxy_url=proxy_url,
            cert_reqs='CERT_NONE',
            assert_hostname=False,
            headers={
                "Proxy-Authorization": f"Basic {auth}",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept": "application/json",
                "Connection": "keep-alive"
            }
        )
        
        response = http.request("GET", "https://ip.decodo.com/json", timeout=15.0)
        
        if response.status == 200:
            data = json.loads(response.data.decode())
            log_success(f"    ✅ urllib3 OK")
            return True, data, "urllib3"
        else:
            log_debug(f"    urllib3 falló con status {response.status}")
            
    except Exception as e:
        log_debug(f"    urllib3 falló: {str(e)[:100]}")
    
    return False, None, "urllib3_none"

def test_proxy_requests(country: str = "ar", username_variant: Optional[str] = None) -> Tuple[bool, Optional[Dict], str]:
    """Prueba el proxy usando requests."""
    try:
        proxy = decodo_build_proxy(country=country, username_variant=username_variant)
        proxy_url = f"http://{proxy['username']}:{proxy['password']}@{proxy['server'].replace('http://', '')}"
        
        log_debug(f"  Probando requests con username: {proxy['username']}")
        
        proxies = {"http": proxy_url, "https": proxy_url}
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "application/json",
            "Proxy-Connection": "Keep-Alive"
        }
        
        session = requests.Session()
        session.trust_env = False
        
        response = session.get(
            "https://ip.decodo.com/json",
            proxies=proxies,
            headers=headers,
            timeout=15,
            verify=False
        )
        
        if response.status_code == 200:
            data = response.json()
            log_success(f"    ✅ requests OK")
            return True, data, "requests"
        else:
            log_debug(f"    requests falló con status {response.status_code}")
            
    except Exception as e:
        log_debug(f"    requests falló: {str(e)[:100]}")
    
    return False, None, "requests_none"

def decodo_test_connection_hybrid(country: str = "ar") -> Tuple[bool, Optional[Dict], str, str]:
    """Prueba híbrida del proxy con múltiples variantes."""
    log_proxy(f"Iniciando test híbrido para {country.upper()}...")
    
    for username_variant in USERNAME_VARIANTS:
        log_proxy(f"\n  Probando variante: {username_variant}")
        
        success, data, method = test_proxy_playwright(country, username_variant)
        if success:
            return True, data, method, username_variant
        
        success, data, method = test_proxy_urllib3(country, username_variant)
        if success:
            return True, data, method, username_variant
        
        success, data, method = test_proxy_requests(country, username_variant)
        if success:
            return True, data, method, username_variant
    
    return False, None, "none", "none"

# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def money_str_to_float(s: str) -> float:
    """Convierte string de precio ARS a float."""
    if s is None:
        return 0.0
    raw = str(s).replace("\xa0", " ").strip()
    if not raw:
        return 0.0
    
    log_debug(f"  Parseando precio: '{raw}'")
    
    cleaned = re.sub(r"[^0-9\.,]", "", raw)
    if not cleaned:
        return 0.0
    
    try:
        if "," in cleaned:
            left, right = cleaned.rsplit(",", 1)
            if 1 <= len(right) <= 2:
                left_digits = left.replace(".", "")
                right_digits = right.ljust(2, "0")[:2]
                val = float(f"{left_digits}.{right_digits}")
            else:
                val = float(cleaned.replace(",", "").replace(".", ""))
        else:
            if re.fullmatch(r"\d{1,3}(?:\.\d{3})+\d{2}", cleaned):
                digits = cleaned.replace(".", "")
                val = float(f"{digits[:-2]}.{digits[-2:]}")
            else:
                val = float(cleaned.replace(".", ""))
        
        while val > 1_000_000:
            val = val / 100.0
        
        log_debug(f"  → Precio parseado: {val:.2f}")
        return float(val)
    except Exception as e:
        log_debug(f"  ⚠️ Error parseando: {e}")
        return 0.0

def safe_round_money(x: float) -> float:
    if x is None:
        return 0.0
    try:
        v = float(x)
        while v > 1_000_000:
            v = v / 100.0
        return float(round(v))
    except Exception:
        return 0.0

def normalize_mla(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    m = re.search(r"(\d{6,})", s)
    if not m:
        return ""
    return f"MLA-{m.group(1)}"

def now_ts() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")

def _compute_off(list_price: float, final_price: float) -> float:
    """Calcula markdown como decimal (0.1 = 10%)."""
    try:
        if list_price <= 0 or final_price <= 0:
            return 0.0
        return max(0.0, (1.0 - (final_price / list_price)))
    except Exception:
        return 0.0

def calc_nike_cuotas(nike_final_ars: float) -> int:
    """Calcula cuotas Nike según modelo."""
    try:
        p = float(nike_final_ars or 0)
    except Exception:
        p = 0.0
    if not NIKE_CUOTAS_SIMPLE_MODE:
        return int(NIKE_CUOTAS_ALL)
    return int(NIKE_CUOTAS_HIGH if p >= NIKE_CUOTAS_HIGH_FROM_ARS else NIKE_CUOTAS_ALL)


def _bml_label(comp: float, nike: float) -> str:
    """
    Beat/Meet/Lose según tu definición (Nike-centric):

    - Beat  => Nike es MÁS BARATO que el competidor (Nike < Comp)
    - Lose  => Nike es MÁS CARO que el competidor (Nike > Comp)
    - Meet  => Diferencia dentro de ±2%

    Nota: trabaja con precios finales (o con precios + shipping cuando se reutiliza).
    """
    try:
        comp = float(comp)
        nike = float(nike)
        if comp <= 0 or nike <= 0:
            return ""
        diff = (nike - comp) / comp  # positivo = Nike más caro, negativo = Nike más barato
        if abs(diff) <= BML_THRESHOLD_PCT:
            return "Meet"
        return "Beat" if diff < -BML_THRESHOLD_PCT else "Lose"
    except Exception:
        return ""


def _bml_full_price(comp_full: float, nike_full: float) -> str:
    """
    Beat/Meet/Lose para precio full (mismo criterio Nike-centric):

    - Beat  => Nike FULL es MÁS BARATO que el full del competidor
    - Lose  => Nike FULL es MÁS CARO que el full del competidor
    - Meet  => Diferencia dentro de ±2%
    """
    try:
        comp_full = float(comp_full)
        nike_full = float(nike_full)
        if comp_full <= 0 or nike_full <= 0:
            return ""
        diff = (nike_full - comp_full) / comp_full
        if abs(diff) <= BML_THRESHOLD_PCT:
            return "Meet"
        return "Beat" if diff < -BML_THRESHOLD_PCT else "Lose"
    except Exception:
        return ""

def _nike_final_from_sale(nike_full: float, sale_str: str) -> float:
    """
    Calcula precio final Nike a partir de markdown decimal (0.1 = 10%).
    Si el string es '10%' o '0.1', lo interpreta correctamente.
    """
    try:
        nike_full = float(nike_full)
        if nike_full <= 0:
            return 0.0
        s = (sale_str or "").strip()
        if not s:
            return nike_full
        # Si es porcentaje tipo '10%' o '10.0%'
        m_pct = re.search(r"(\d+(?:[.,]\d+)?)%", s)
        if m_pct:
            pct = float(m_pct.group(1).replace(",", "."))
            pct = max(0.0, min(100.0, pct))
            return nike_full * (1.0 - pct/100.0)
        # Si es decimal tipo '0.1' o '0,1'
        m_dec = re.search(r"(0[.,]\d+)", s)
        if m_dec:
            dec = float(m_dec.group(1).replace(",", "."))
            dec = max(0.0, min(1.0, dec))
            return nike_full * (1.0 - dec)
        # Si es solo número, asume porcentaje
        m = re.search(r"(\d+(?:[.,]\d+)?)", s)
        if m:
            pct = float(m.group(1).replace(",", "."))
            if pct > 1.0:
                pct = max(0.0, min(100.0, pct))
                return nike_full * (1.0 - pct/100.0)
            else:
                dec = max(0.0, min(1.0, pct))
                return nike_full * (1.0 - dec)
        return nike_full
    except Exception:
        return float(nike_full or 0.0)

# ============================================================
# CACHE MANAGEMENT - TTL de 4 días
# ============================================================

class Cache:
    """Sistema de cache con TTL de 4 días."""
    
    def __init__(self, cache_file: str, ttl_days: int = CACHE_TTL_DAYS):
        self.cache_file = Path(cache_file)
        self.ttl = timedelta(days=ttl_days)
        self.cache: Dict[str, Any] = {}
        self.load()
    
    def load(self) -> None:
        """Carga cache desde archivo, filtrando entradas expiradas."""
        if not self.cache_file.exists():
            log_info(f"📁 Cache nuevo: {self.cache_file.name}")
            return
        
        try:
            data = json.loads(self.cache_file.read_text(encoding="utf-8"))
            now = datetime.now()
            valid = 0
            expired = 0
            
            for key, value in data.items():
                if "last_updated" in value:
                    try:
                        last_upd = datetime.fromisoformat(value["last_updated"])
                        if now - last_upd < self.ttl:
                            self.cache[key] = value
                            valid += 1
                        else:
                            expired += 1
                            log_debug(f"  ⌛ Entrada expirada: {key} (última actualización: {last_upd.strftime('%Y-%m-%d')})")
                    except:
                        self.cache[key] = value
                        valid += 1
                else:
                    self.cache[key] = value
                    valid += 1
            
            log_info(f"📦 Cache cargado: {valid} válidas, {expired} expiradas (TTL: {CACHE_TTL_DAYS} días)")
            
        except Exception as e:
            log_warning(f"⚠️ Error cargando cache: {e}")
    
    def save(self) -> None:
        """Guarda cache a archivo."""
        try:
            self.cache_file.write_text(
                json.dumps(self.cache, ensure_ascii=False, indent=2, default=str),
                encoding="utf-8"
            )
            log_debug(f"💾 Cache guardado: {self.cache_file.name} ({len(self.cache)} entries)")
        except Exception as e:
            log_warning(f"⚠️ Error guardando cache: {e}")
    
    def get(self, key: str) -> Optional[Any]:
        """Obtiene valor del cache si no está expirado."""
        if key in self.cache:
            value = self.cache[key]
            if "last_updated" in value:
                try:
                    last_upd = datetime.fromisoformat(value["last_updated"])
                    if datetime.now() - last_upd < self.ttl:
                        return value
                    else:
                        log_debug(f"  ⌛ Entrada expirada en get: {key}")
                        del self.cache[key]
                except:
                    return value
            return value
        return None
    
    def set(self, key: str, value: Any) -> None:
        """Guarda valor en cache con timestamp actual."""
        value["last_updated"] = datetime.now().isoformat()
        self.cache[key] = value
    
    def delete(self, key: str) -> None:
        """Elimina del cache."""
        if key in self.cache:
            del self.cache[key]
    
    def size(self) -> int:
        return len(self.cache)

# ============================================================
# LECTURA DEL STATUSBOOK
# ============================================================

def read_statusbook_universe(xlsb_path: str, season: str) -> List[Dict[str, Any]]:
    """Lee StatusBooks y arma el universo filtrado.

    Reglas "indestructibles" para Nike Full Price (según requerimiento):
    - Nike Full Price Num sale SIEMPRE de la columna Season (ej: SP26).
    - Si hay MÁS DE UNA columna Season (headers duplicados), se toma:
        * por fila: el PRIMER valor válido (>0) de izquierda a derecha entre las columnas duplicadas
        * si no hay ningún valor válido: 0
    - El SeasonPrice se obtiene por cruce Product Code ↔ SSN VTA:
        * se arma un lookup SSN_VTA -> SeasonPrice (tomando el máximo si hay repetidos)
        * para cada fila, si existe lookup para su Product Code, se usa ese valor
          (fallback: SeasonPrice de la propia fila)
    - Se normalizan claves (Product Code / SSN VTA) para evitar mismatches por espacios/casos raros.
    """
    log_info(f"Leyendo StatusBook: {xlsb_path}")

    if not Path(xlsb_path).exists():
        raise FileNotFoundError(f"❌ No existe StatusBooks en: {xlsb_path}")

    wb = open_workbook(xlsb_path)
    if STATUSBOOK_SHEET not in wb.sheets:
        raise RuntimeError(f"❌ No encuentro sheet '{STATUSBOOK_SHEET}'. Sheets: {wb.sheets}")

    season = (season or "").strip()
    if not season:
        raise ValueError("❌ season vacío.")

    # --------- 1) Leer header y cargar filas ---------
    rows_data: List[List[Any]] = []
    header: Optional[List[Any]] = None

    with wb.get_sheet(STATUSBOOK_SHEET) as sh:
        for i, row in enumerate(sh.rows()):
            vals = [c.v for c in row]
            if i == STATUSBOOK_HEADER_ROW_IDX:
                header = vals
                log_debug(f"Header encontrado en fila {i+1}: {header[:14]}...")
                continue
            if i <= STATUSBOOK_HEADER_ROW_IDX:
                continue
            rows_data.append(vals)

    if not header:
        raise RuntimeError("❌ No pude leer header del StatusBooks.")

    # --------- 2) Mapeo de columnas robusto (soporta duplicados) ---------
    from collections import defaultdict

    def _hnorm(x: Any) -> str:
        if not isinstance(x, str):
            return ""
        return x.strip()

    col_multi: Dict[str, List[int]] = defaultdict(list)
    for idx, h in enumerate(header):
        hn = _hnorm(h)
        if hn:
            col_multi[hn].append(idx)

    def col_first(name: str) -> int:
        name = (name or "").strip()
        if name not in col_multi or not col_multi[name]:
            raise KeyError(name)
        return col_multi[name][0]

    def col_all(name: str) -> List[int]:
        name = (name or "").strip()
        if name not in col_multi or not col_multi[name]:
            raise KeyError(name)
        return list(col_multi[name])

    required = [
        "Product Code", "SSN VTA", "Style", "MLA", "Marketing Name", "BU",
        "Category", "Gender", "Franchise", "SALE",
        "STOCK BL (Inventario Brandlive)", "MELI"
    ]
    missing = [r for r in required if (r.strip() not in col_multi)]
    if season not in col_multi:
        missing.append(season)
    if missing:
        raise RuntimeError(f"❌ Faltan columnas: {missing}")

    idx_product = col_first("Product Code")
    idx_ssn = col_first("SSN VTA")
    idx_style = col_first("Style")
    idx_mla = col_first("MLA")
    idx_mkt = col_first("Marketing Name")
    idx_bu = col_first("BU")
    idx_cat = col_first("Category")
    idx_gen = col_first("Gender")
    idx_fra = col_first("Franchise")
    idx_sale = col_first("SALE")
    idx_stock = col_first("STOCK BL (Inventario Brandlive)")
    idx_meli = col_first("MELI")

    season_idxs = col_all(season)  # <- TODAS las columnas "SP26" (si hay duplicados)

    log_debug(f"Columnas únicas mapeadas: {len(col_multi)} | Season '{season}' ocurrencias: {len(season_idxs)}")

    # --------- 3) Helpers de parsing / normalización ---------
    def norm_key(s: Any) -> str:
        # Normalización defensiva (mismatches por espacios / casing / NBSP)
        if s is None:
            return ""
        t = str(s)
        t = t.replace("\u00A0", " ")  # NBSP
        t = t.strip().upper()
        t = re.sub(r"\s+", "", t)     # sin espacios
        return t

    def parse_float_safe(v: Any) -> float:
        try:
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                if math.isnan(float(v)):
                    return 0.0
                return float(v)
            s = str(v).strip()
            if not s:
                return 0.0
            # Algunas celdas pueden venir como "229.999" o "229,999.00"
            return money_str_to_float(s) or 0.0
        except Exception:
            return 0.0

    def season_value_for_row(vals: List[Any]) -> float:
        # PRIMER valor válido (>0) entre las columnas duplicadas del season (izq->der)
        for j in season_idxs:
            if j >= len(vals):
                continue
            fv = parse_float_safe(vals[j])
            if fv > 0:
                return fv
        return 0.0

    # --------- 4) Lookup SSN VTA -> SeasonPrice (máximo si hay repetidos) ---------
    ssn_to_season: Dict[str, float] = {}
    for vals in rows_data:
        ssn_key = norm_key(vals[idx_ssn])
        if not ssn_key:
            continue
        sp = season_value_for_row(vals)
        if sp > 0:
            prev = ssn_to_season.get(ssn_key, 0.0)
            if sp > prev:
                ssn_to_season[ssn_key] = sp

    # --------- 5) Construir universo filtrado (dedupe por StyleColor) ---------
    # Si el StatusBooks tiene filas duplicadas para el mismo Product Code,
    # nos quedamos con la "mejor" según: SeasonPrice (desc), StockBL (desc).
    best_by_style: Dict[str, Dict[str, Any]] = {}
    total_rows = 0

    for vals in rows_data:
        total_rows += 1

        product_raw = vals[idx_product]
        product_key = norm_key(product_raw)
        if not product_key:
            continue

        # SeasonPrice por cruce: Product Code busca en SSN VTA
        season_price_row = season_value_for_row(vals)
        season_price = ssn_to_season.get(product_key, season_price_row)
        if season_price <= 0:
            continue

        stock_bl = parse_float_safe(vals[idx_stock])
        if stock_bl <= 0:
            continue

        meli_flag = str(vals[idx_meli] or "").strip().upper()
        if meli_flag not in ("MELI", "SI", "SÍ"):
            continue

        mla = normalize_mla(vals[idx_mla])
        if not mla:
            continue

        # Nike Full = SIEMPRE SeasonPrice (indestructible)
        full_price = float(season_price)

        sale_str = str(vals[idx_sale] or "").strip()
        nike_final = _nike_final_from_sale(full_price, sale_str)

        row_obj = {
            "StyleColor": str(product_raw or "").strip(),
            "ProductCode": str(product_raw or "").strip(),
            "SSN VTA": str(vals[idx_ssn] or "").strip(),
            "Style": str(vals[idx_style] or "").strip(),
            "Marketing Name": str(vals[idx_mkt] or "").strip(),
            "Division": str(vals[idx_bu] or "").strip(),
            "Category": str(vals[idx_cat] or "").strip(),
            "Gender": str(vals[idx_gen] or "").strip(),
            "Franchise": str(vals[idx_fra] or "").strip(),
            "MLA": mla,
            "Nike Full Price Num": float(full_price),
            "Nike Markdown Str": sale_str,
            "Nike Final Price Num": float(nike_final),
            "Stock BL": float(stock_bl),
        }

        # Dedupe
        prev = best_by_style.get(product_key)
        if prev is None:
            best_by_style[product_key] = row_obj
        else:
            # criterio: SeasonPrice desc, luego Stock desc
            prev_full = float(prev.get("Nike Full Price Num", 0.0) or 0.0)
            prev_stock = float(prev.get("Stock BL", 0.0) or 0.0)
            if (full_price > prev_full) or (full_price == prev_full and stock_bl > prev_stock):
                best_by_style[product_key] = row_obj

    universe = list(best_by_style.values())
    log_info(f"Universo filtrado: {len(universe)} productos únicos (dedupe) | {total_rows} filas totales")
    return universe


NON_ACTIVE_PHRASES = [
    'publicación pausada',
    'publicacion pausada',
    'publicación finalizada',
    'publicacion finalizada',
    'pausada temporalmente',
]

def page_has_nonactive_marker(page) -> Optional[str]:
    """Detecta si la página está pausada/finalizada."""
    try:
        if page.get_by_text('Publicación pausada', exact=False).count() > 0:
            return 'publicación pausada'
        if page.get_by_text('Publicación finalizada', exact=False).count() > 0:
            return 'publicación finalizada'
        if page.get_by_text('Pausada temporalmente', exact=False).count() > 0:
            return 'pausada temporalmente'
    except Exception:
        pass
    
    try:
        txt = (page.inner_text('body', timeout=5_000) or '').lower()
    except Exception:
        return None
    for p in NON_ACTIVE_PHRASES:
        if p in txt:
            return p
    return None



def meli_is_buyable(page) -> bool:
    """Heurística robusta: verifica si la publicación es 'comprable' en MeLi (solo PDP).
    No alcanza con HTTP 200: muchas publicaciones muestran precio pero no permiten comprar.
    Criterios:
    - Debe existir un CTA de compra típico (Comprar ahora / Agregar al carrito) y NO estar disabled.
    - Si aparecen textos claros de 'sin stock' / 'no disponible', devuelve False.
    """
    try:
        body_txt = (page.inner_text("body", timeout=2_000) or "").lower()
        # señales fuertes de no comprable
        for bad in [
            "sin stock", "no disponible", "producto no disponible", "publicación pausada",
            "publicacion pausada", "publicación finalizada", "publicacion finalizada",
            "no está disponible", "no esta disponible"
        ]:
            if bad in body_txt:
                return False
    except Exception:
        body_txt = ""

    # CTA típicos (MeLi cambia clases, por eso buscamos por texto y por contenedores comunes)
    cta_locators = [
        'button:has-text("Comprar")',
        'a:has-text("Comprar")',
        'button:has-text("Agregar al carrito")',
        'a:has-text("Agregar al carrito")',
        'button:has-text("Comprar ahora")',
        'a:has-text("Comprar ahora")',
        'button.ui-pdp-action-primary',
        'a.ui-pdp-action-primary',
        'button.ui-pdp-buybox__action',
        'a.ui-pdp-buybox__action',
    ]

    try:
        for sel in cta_locators:
            loc = page.locator(sel).first
            if loc.count() == 0:
                continue
            # si está visible y no disabled -> comprable
            try:
                if not loc.is_visible():
                    continue
            except Exception:
                pass
            try:
                disabled = loc.get_attribute("disabled")
                aria_disabled = loc.get_attribute("aria-disabled")
                if (disabled is not None) or (aria_disabled in ("true", "1")):
                    continue
            except Exception:
                pass
            return True
    except Exception:
        pass

    return False

def extract_final_price_from_meta(page) -> Optional[float]:
    """Extrae precio del meta itemprop='price'."""
    try:
        el = page.locator('meta[itemprop="price"]').first
        if el.count() == 0:
            return None
        content = el.get_attribute("content")
        if not content:
            return None
        v = money_str_to_float(content)
        return v if v > 0 else None
    except Exception:
        return None

def extract_off_pct_from_badge(page) -> Optional[float]:
    """Extrae porcentaje de descuento del badge 'X% OFF'."""
    try:
        candidates = page.locator('span:has-text("% OFF")')
        n = candidates.count()
        for i in range(min(n, 8)):
            t = candidates.nth(i).inner_text().strip()
            m = re.search(r"(\d+(?:[.,]\d+)?)\s*%?\s*OFF", t, re.I)
            if m:
                pct = float(m.group(1).replace(",", "."))
                if 0 < pct <= 90:
                    return pct
    except Exception:
        pass
    return None

def extract_preloaded_state_json(page) -> Optional[dict]:
    """Extrae el JSON de __PRELOADED_STATE__."""
    try:
        s = page.locator("script#__PRELOADED_STATE__").first
        if s.count() == 0:
            return None
        raw = s.inner_text().strip()
        if not raw:
            return None
        return json.loads(raw)
    except Exception:
        return None

def walk_find_numbers(obj, hits: List[Tuple[str, float]], path=""):
    """Recorre objeto buscando números."""
    try:
        if isinstance(obj, dict):
            for k, v in obj.items():
                p = f"{path}.{k}" if path else str(k)
                if isinstance(v, (int, float)) and 1 <= float(v) <= 9e7:
                    hits.append((p, float(v)))
                else:
                    walk_find_numbers(v, hits, p)
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                p = f"{path}[{i}]"
                walk_find_numbers(v, hits, p)
    except Exception:
        return

def extract_final_from_state(state: dict) -> Optional[float]:
    """Extrae precio final del preloaded state."""
    if not isinstance(state, dict):
        return None

    probes = [
        ("initialState.components.price", ["amount", "current_price", "value"]),
        ("initialState.components.price_part", ["amount", "current_price", "value"]),
        ("initialState.items[0].price", ["amount", "value"]),
        ("initialState.item.price", ["amount", "value"]),
        ("initialState.components.price.price", ["amount", "value"]),
    ]

    def get_path(root, path_str):
        cur = root
        for part in path_str.split("."):
            if not part:
                continue
            m = re.match(r"(\w+)\[(\d+)\]$", part)
            if m:
                key = m.group(1); idx = int(m.group(2))
                if not isinstance(cur, dict) or key not in cur:
                    return None
                cur = cur[key]
                if not isinstance(cur, list) or idx >= len(cur):
                    return None
                cur = cur[idx]
            else:
                if not isinstance(cur, dict) or part not in cur:
                    return None
                cur = cur[part]
        return cur

    for base_path, keys in probes:
        node = get_path(state, base_path)
        if isinstance(node, dict):
            for k in keys:
                if k in node:
                    v = node[k]
                    if isinstance(v, (int, float)) and float(v) > 0:
                        return float(v)
                    if isinstance(v, dict) and "amount" in v and isinstance(v["amount"], (int, float)):
                        return float(v["amount"])
        if isinstance(node, (int, float)) and float(node) > 0:
            return float(node)

    hits = []
    walk_find_numbers(state, hits)
    if hits:
        price_hits = [(p, v) for p, v in hits if 'price' in p.lower()]
        if price_hits:
            return float(price_hits[0][1]) if price_hits[0][1] > 0 else None
    
    return None

def extract_list_price_from_dom(page, final_num: float) -> Optional[float]:
    """Intenta extraer el precio original/tachado (full/list) desde el DOM.

    Estrategia:
    - Buscar selectores conocidos del "precio original" y dentro de ellos leer andes-money-amount__fraction.
    - Validar que el número sea > final_num (si no, ignorar).
    """
    try:
        final_num = float(final_num or 0.0)
    except Exception:
        final_num = 0.0

    selectors = [
        # Clásicos MeLi PDP
        "span.ui-pdp-price__original-value span.andes-money-amount__fraction",
        "s span.andes-money-amount__fraction",
        ".ui-pdp-price__original-value span.andes-money-amount__fraction",
        ".ui-pdp-price__original-value .andes-money-amount__fraction",
        ".ui-pdp-price__original-value",
        # Variantes / wrappers
        "[data-testid='price-part'] s span.andes-money-amount__fraction",
        "[data-testid='price-part'] .ui-pdp-price__original-value span.andes-money-amount__fraction",
        "span.price-part s span.andes-money-amount__fraction",
    ]

    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count():
                txt = loc.inner_text().strip()
                num = money_str_to_float(txt)
                if num and num > 0:
                    num = safe_round_money(num)
                    if final_num > 0 and num <= final_num:
                        continue
                    return num
        except Exception:
            continue

    return None


def extract_list_from_state(state: dict, final_num: float) -> Optional[float]:
    """Intenta extraer precio original/list desde preloaded state (JS).

    Busca claves típicas: original / before_discount / previous / regular.
    """
    if not isinstance(state, dict):
        return None

    try:
        final_num = float(final_num or 0.0)
    except Exception:
        final_num = 0.0

    # Heurística por paths comunes
    candidate_paths = [
        "initialState.components.price.original_amount",
        "initialState.components.price.original_price",
        "initialState.components.price.previous_amount",
        "initialState.components.price.before_discount_amount",
        "initialState.components.price_part.original_amount",
        "initialState.components.price_part.original_price",
        "initialState.item.original_price",
        "initialState.item.price_before_discount",
        "initialState.items[0].original_price",
        "initialState.items[0].price_before_discount",
    ]

    def get_path(root, path_str):
        cur = root
        for part in path_str.split("."):
            if not part:
                continue
            m = re.match(r"(\w+)\[(\d+)\]$", part)
            if m:
                key = m.group(1); idx = int(m.group(2))
                if not isinstance(cur, dict) or key not in cur:
                    return None
                cur = cur[key]
                if not isinstance(cur, list) or idx >= len(cur):
                    return None
                cur = cur[idx]
            else:
                if not isinstance(cur, dict) or part not in cur:
                    return None
                cur = cur[part]
        return cur

    for pth in candidate_paths:
        node = get_path(state, pth)
        if isinstance(node, (int, float)) and float(node) > 0:
            v = float(node)
            if final_num > 0 and v <= final_num:
                continue
            return safe_round_money(v)
        if isinstance(node, dict):
            # a veces {amount: x}
            for k in ("amount", "value"):
                if k in node and isinstance(node[k], (int, float)):
                    v = float(node[k])
                    if final_num > 0 and v <= final_num:
                        continue
                    return safe_round_money(v)

    # Heurística global: buscar números con "original"/"regular"/"previous" en el path
    hits: List[Tuple[str, float]] = []
    walk_find_numbers(state, hits)
    if hits:
        prefer = []
        for p, v in hits:
            pl = p.lower()
            if any(k in pl for k in ("original", "before", "previous", "regular", "was", "list")):
                if v > 0 and (final_num <= 0 or v > final_num):
                    prefer.append((p, v))
        if prefer:
            # elegir el menor > final (más razonable)
            prefer_sorted = sorted(prefer, key=lambda x: x[1])
            return safe_round_money(prefer_sorted[0][1])

    return None


def extract_list_price_fallback(page, final_num: float) -> Optional[float]:
    """Fallback ultra-robusto:
    - Recolecta montos cercanos al price card.
    - Devuelve el menor monto que sea > final_num.
    """
    try:
        final_num = float(final_num or 0.0)
    except Exception:
        final_num = 0.0

    candidates: List[float] = []

    containers = [
        "div.ui-pdp-price",
        "div.ui-pdp-price__main-container",
        "[data-testid='price-part']",
        "section.ui-pdp-price",
    ]

    for cont in containers:
        try:
            root = page.locator(cont).first
            if not root.count():
                continue
            fracs = root.locator("span.andes-money-amount__fraction")
            n = fracs.count()
            for i in range(min(n, 30)):
                try:
                    t = fracs.nth(i).inner_text().strip()
                    v = money_str_to_float(t)
                    if v and v > 0:
                        candidates.append(safe_round_money(v))
                except Exception:
                    continue
        except Exception:
            continue

    # Si no encontré en contenedores, pruebo global limitado (último recurso)
    if not candidates:
        try:
            fracs = page.locator("span.andes-money-amount__fraction")
            n = fracs.count()
            for i in range(min(n, 25)):
                try:
                    t = fracs.nth(i).inner_text().strip()
                    v = money_str_to_float(t)
                    if v and v > 0:
                        candidates.append(safe_round_money(v))
                except Exception:
                    continue
        except Exception:
            pass

    if not candidates:
        return None

    # normalizar y elegir menor > final
    if final_num > 0:
        bigger = sorted({c for c in candidates if c > final_num})
        if bigger:
            return bigger[0]
        return None

    # si no tengo final, devolver el máximo (pero esto casi no debería pasar)
    return max(candidates) if candidates else None


def resolve_meli_list_price(page, final_num: float) -> Optional[float]:
    """Orquestador del list/full price de MeLi: DOM -> State -> Fallback numérico."""
    # 1) DOM (precio original/tachado)
    dom_val = extract_list_price_from_dom(page, final_num)
    if dom_val and dom_val > 0:
        return dom_val

    # 2) JS state
    try:
        state = extract_preloaded_state_json(page)
    except Exception:
        state = None
    if state:
        st_val = extract_list_from_state(state, final_num)
        if st_val and st_val > 0:
            return st_val

    # 3) fallback por candidatos cerca del price card
    fb_val = extract_list_price_fallback(page, final_num)
    if fb_val and fb_val > 0:
        return fb_val

    return None

def extract_title(page) -> str:
    """Extrae título del producto."""
    for sel in ["h1.ui-pdp-title", "h1"]:
        try:
            loc = page.locator(sel).first
            if loc.count():
                t = loc.inner_text().strip()
                if t:
                    return t[:200]
        except Exception:
            pass
    return ""

def extract_cuotas(page, final_price: float) -> int:
    """Extrae cuotas sin interés."""
    try:
        body = page.inner_text("body", timeout=5_000)
    except Exception:
        return 0
    text = " ".join(body.split())
    text_low = text.lower()

    if "cuota" not in text_low:
        return 0

    best = 0
    for m in re.finditer(r"(\d{1,2})\s*cuotas?\s*(?:sin\s+inter[eé]s|sin\s+interes)", text_low):
        try:
            q = int(m.group(1))
            best = max(best, q)
        except Exception:
            pass

    for m in re.finditer(r"(\d{1,2})\s*cuotas?\s*de\s*\$?\s*([\d\.]+(?:,\d{1,2})?)", text_low):
        try:
            q = int(m.group(1))
            amt = money_str_to_float(m.group(2))
            if q <= 0 or amt <= 0:
                continue
            total = q * amt
            if final_price > 0 and abs(total - final_price) / final_price <= CUOTAS_TOL_PCT:
                best = max(best, q)
        except Exception:
            pass

    return int(best)

def is_retryable_error(error_str: str) -> bool:
    """Determina si un error es reintentable."""
    error_lower = error_str.lower()
    for pattern in RETRY_ON_ERRORS:
        if pattern.lower() in error_lower:
            return True
    return False

def scrape_meli_pdp_with_retry(page, mla: str, nike_full: float, attempt: int = 1) -> Dict[str, Any]:
    """
    Scrapea PDP con reintentos para errores transitorios.
    """
    import time  # Garantiza disponibilidad en multiprocessing
    log_scraping(f"Procesando MLA: {mla} (intento {attempt}/{MAX_SCRAPE_RETRIES})")
    
    out = {
        "ok": False,
        "status": "init",
        "url": f"https://articulo.mercadolibre.com.ar/{mla}",
        "title": "",
        "final_price_num": None,
        "list_price_num": None,
        "off_pct": None,
        "cuotas_sin_interes": 0,
        "is_active": False,
        "is_buyable": False,
        "http_status": None,
        "error_detail": None,
        "attempts": attempt,
        "needs_rebuild": False
    }

    url = out["url"]
    try:
        log_debug(f"  Navegando a {url}")
        response = page.goto(url, wait_until="domcontentloaded", timeout=DEFAULT_NAV_TIMEOUT_MS)
        status_code = None
        try:
            status_code = response.status if response else None
        except:
            pass
        out["http_status"] = status_code
        log_http(f"Status code: {status_code}", status_code)
        # Advanced retry and proxy rotation for 403/429
        if status_code in [403, 429]:
            out["status"] = "proxy_block"
            out["error_detail"] = f"Proxy/IP bloqueado ({status_code})"
            out["needs_rebuild"] = True
            log_warning(f"  ⚠️ Proxy/IP bloqueado ({status_code}) para {mla}, rotando IP...")
            if attempt < MAX_SCRAPE_RETRIES:
                # Rotate IP and retry
                from uuid import uuid4
                rotate_ip_func = globals().get("rotate_ip")
                if rotate_ip_func:
                    # This assumes you have access to browser/context/page/playwright in scope
                    log_debug("  Rotando IP por 403/429...")
                    # NOTE: Actual rotation must be handled in worker_process_chunk
                wait_time = RETRY_BACKOFF_FACTOR ** attempt
                log_warning(f"  ⏳ Reintentando en {wait_time}s...")
                import time
                time.sleep(wait_time)
                return scrape_meli_pdp_with_retry(page, mla, nike_full, attempt + 1)
            return out
        if status_code and status_code >= 400:
            if status_code == 404:
                out["status"] = "not_found"
                out["error_detail"] = "Página no encontrada (404)"
                log_warning(f"  ⚠️ 404 Not Found para {mla}")
                return out
            elif status_code >= 500:
                out["status"] = "server_error"
                out["error_detail"] = f"Error del servidor ({status_code})"
                log_warning(f"  ⚠️ {status_code} Server Error para {mla}")
                return out
        page.wait_for_timeout(800)
        cur_url = page.url or ""
        if ("account-verification" in cur_url) or ("/gz/" in cur_url and "verification" in cur_url) or ("captcha" in cur_url.lower()):
            out["status"] = "login_required"
            log_warning("  ⚠️ Login/CAPTCHA detectado")
            return out
        reason = page_has_nonactive_marker(page)
        if reason:
            out["status"] = f"non_active:{reason}"
            log_warning(f"  ⚠️ Producto no activo: {reason}")
            return out
        out["title"] = extract_title(page)
        # Disponibilidad real (buybox) para evitar falsos BML Beat cuando no está comprable
        out["is_buyable"] = meli_is_buyable(page)
        # Expanded price extraction logic
        price_log_details = []
        final_num = extract_final_price_from_meta(page)
        price_log_details.append(f"extract_final_price_from_meta={final_num}")
        if final_num is None:
            state = extract_preloaded_state_json(page)
            price_log_details.append(f"extract_preloaded_state_json={bool(state)}")
            if state:
                final_num = extract_final_from_state(state)
                price_log_details.append(f"extract_final_from_state={final_num}")
        if final_num is None:
            try:
                frac = page.locator("span.andes-money-amount__fraction").first
                if frac.count():
                    final_num = money_str_to_float(frac.inner_text())
                    price_log_details.append(f"andes-money-amount__fraction={final_num}")
            except Exception as e:
                price_log_details.append(f"andes-money-amount__fraction error: {e}")
        # Try more selectors
        if final_num is None:
            try:
                val = page.locator("span.ui-pdp-price__second-line").first
                if val.count():
                    final_num = money_str_to_float(val.inner_text())
                    price_log_details.append(f"ui-pdp-price__second-line={final_num}")
            except Exception as e:
                price_log_details.append(f"ui-pdp-price__second-line error: {e}")
        if final_num is None or final_num <= 0:
            out["status"] = "no_final_price"
            log_warning(f"  ⚠️ No se pudo extraer precio final | Detalles: {' | '.join(price_log_details)}")
            out["error_detail"] = ' | '.join(price_log_details)
            return out
        final_num = safe_round_money(final_num)
        log_success(f"  💰 Precio final: ${final_num:,.2f}")
        list_num = resolve_meli_list_price(page, final_num)
        if list_num is None or float(list_num) <= 0:
            list_num = final_num
        off_badge = extract_off_pct_from_badge(page)
        if off_badge is not None and off_badge > 0:
            off_pct = float(off_badge)
        else:
            off_pct = _compute_off(list_num, final_num)
        cuotas = extract_cuotas(page, final_num)
        out.update({
            "ok": True,
            "status": "ok",
            "final_price_num": float(final_num),
            "list_price_num": float(list_num),
            "off_pct": float(off_pct),
            "cuotas_sin_interes": int(cuotas),
            "is_active": True,
            "price_log_details": price_log_details
        })
        log_success(f"  ✅ PDP scrapeada exitosamente (intento {attempt})")
        return out

    except PWTimeoutError:
        out["status"] = "timeout"
        out["error_detail"] = "Timeout"
        log_warning(f"  ⏰ Timeout cargando {mla}")
        
        if attempt < MAX_SCRAPE_RETRIES:
            wait_time = RETRY_BACKOFF_FACTOR ** attempt
            log_warning(f"  ⏳ Reintentando en {wait_time}s...")
            time.sleep(wait_time)
            return scrape_meli_pdp_with_retry(page, mla, nike_full, attempt + 1)
        return out
        
    except PWError as e:
        error_str = str(e)
        # Si se rompe el túnel del proxy, esto NO se arregla con retry sobre el mismo page.
        # Marcamos needs_rebuild para que el worker cierre y reconstruya browser/context.
        if any(pat.lower() in error_str.lower() for pat in TUNNEL_ERROR_PATTERNS):
            out["status"] = "tunnel_failed"
            out["needs_rebuild"] = True
        else:
            out["status"] = "playwright_error"

        out["error_detail"] = error_str[:200]
        
        if "ERR_HTTP_RESPONSE_CODE_FAILURE" in error_str:
            log_http(f"  ⚠️ Error HTTP en {mla}", 400)
        else:
            log_error(f"  ❌ Error de Playwright: {error_str[:100]}")
        
        # Reintentar si es un error reintentable
        if is_retryable_error(error_str) and attempt < MAX_SCRAPE_RETRIES:
            wait_time = RETRY_BACKOFF_FACTOR ** attempt
            log_warning(f"  ⏳ Error reintentable, reintentando en {wait_time}s...")
            time.sleep(wait_time)
            return scrape_meli_pdp_with_retry(page, mla, nike_full, attempt + 1)
        
        return out
        
    except Exception as e:
        log_error(f"  ❌ Error inesperado: {e}")
        out["status"] = f"error:{str(e)[:120]}"
        
        if attempt < MAX_SCRAPE_RETRIES:
            wait_time = RETRY_BACKOFF_FACTOR ** attempt
            log_warning(f"  ⏳ Reintentando en {wait_time}s...")
            time.sleep(wait_time)
            return scrape_meli_pdp_with_retry(page, mla, nike_full, attempt + 1)
        
        return out

# ============================================================
# CONSTRUCCIÓN DE OUTPUT
# ============================================================

OUTPUT_COLS = [
    "StyleColor", "ProductCode", "Marketing Name", "Category", "Division", "Franchise", "Gender",
    "Link PDP Competitor", "Competitor Full Price", "Competitor Markdown", "Competitor Final Price",
    "Nike Full Price", "Nike Markdown", "Nike Final Price", "Competitor vs Nike",
    "BML Final Price", "BML Full Price", "Competitor Shipping", "Nike Shipping",
    "Nike Price + Shipping", "Competitor Price + Shipping", "BML with Shipping",
    "Cuotas Competitor", "Cuotas Nike", "BML Cuotas", "Competitor", "Season", "Fecha Corrida", "Last Update Competitor",
]

def format_money(v: Any) -> Any:
    """Formatea números como moneda ARS ($ XX.XXX)."""
    try:
        if v is None:
            return ""
        v = float(v)
        if v <= 0:
            return ""
        s = f"{int(round(v)):,}".replace(",", ".")
        return f"$ {s}"
    except Exception:
        return ""

def format_md(v: Any) -> str:
    """Formatea markdown como decimal (0.1 = 10%)."""
    try:
        if v is None:
            return ""
        v = float(v)
        if v <= 0:
            return "0.0"
        return f"{v:.3f}"
    except Exception:
        return ""

def build_output_row(base: Dict[str, Any], meli: Dict[str, Any], season: str, run_ts: str) -> Dict[str, Any]:
    """Construye una fila de output según especificación."""
    nike_full = float(base.get("Nike Full Price Num") or 0.0)
    nike_final = float(base.get("Nike Final Price Num") or 0.0)

    # Season en output: valor de SSN VTA por producto (si existe)
    output_season = str(base.get("SSN VTA") or season)

    if nike_full <= 0:
        nike_full = nike_final

    # Markdown de Nike como decimal
    nike_md = _compute_off(nike_full, nike_final)
    nike_md_str = base.get("Nike Markdown Str")
    if nike_md_str:
        # Si viene como porcentaje, convertir a decimal
        m_pct = re.match(r"(\d+(?:[.,]\d+)?)%", nike_md_str)
        if m_pct:
            nike_md = float(m_pct.group(1).replace(",", ".")) / 100.0
        else:
            try:
                nike_md = float(nike_md_str.replace(",", "."))
            except Exception:
                pass

    comp_final = meli.get("final_price_num") if meli else None
    comp_full = None
    if meli:
        comp_full = meli.get("list_price_num")
    if comp_full is None or (isinstance(comp_full, (int, float)) and float(comp_full) <= 0):
        comp_full = comp_final if comp_final is not None else nike_full
    # sanity: si por alguna razón comp_full < comp_final, fallback a comp_final
    try:
        if comp_final is not None and float(comp_full) < float(comp_final):
            comp_full = comp_final
    except Exception:
        pass
    comp_md = _compute_off(comp_full, float(comp_final)) if (comp_final is not None and comp_full > 0) else 0.0

    # Competitor vs Nike (decimal, sin '+')
    comp_vs_nike = ""
    try:
        if comp_final is not None and nike_final > 0:
            diff = (float(comp_final) - float(nike_final)) / float(nike_final)
            comp_vs_nike = f"{diff:.3f}"
    except Exception:
        comp_vs_nike = ""

    # BMLs (Nike más barato = Beat; Nike más caro = Lose)
    bml_final = _bml_label(float(comp_final or 0.0), nike_final)
    # Regla: Final Price no debería dar Beat; si da Beat, validar buybox en MeLi.
    # Si NO está comprable, anulamos BML Final (NA) para evitar falsos positivos por disponibilidad.
    try:
        if bml_final == "Beat":
            is_buyable = bool(meli.get("is_buyable")) if meli else False
            if not is_buyable:
                bml_final = ""
    except Exception:
        pass
    bml_full = _bml_full_price(comp_full, nike_full)

    # Shipping
    comp_ship = float(0 if (comp_final or 0) >= MELI_FREE_SHIP_FROM_ARS else MELI_STD_SHIPPING_ARS)
    nike_ship = float(0 if (nike_final or 0) >= NIKE_FREE_SHIP_FROM_ARS else NIKE_STD_SHIPPING_ARS)

    nike_plus_ship = safe_round_money((nike_final or 0) + nike_ship)
    comp_plus_ship = safe_round_money((comp_final or 0) + comp_ship)

    bml_ship = ""
    if nike_plus_ship > 0 and comp_plus_ship > 0:
        bml_ship = _bml_label(comp_plus_ship, nike_plus_ship)

    # Cuotas
    comp_cuotas = int(meli.get("cuotas_sin_interes") or 0) if meli else 0
    nike_cuotas = int(calc_nike_cuotas(nike_final or 0)) if (nike_final or 0) > 0 else 0

    bml_cuotas = ""
    try:
        if comp_cuotas == nike_cuotas:
            bml_cuotas = "Meet"
        else:
            bml_cuotas = "Beat" if nike_cuotas > comp_cuotas else "Lose"
    except Exception:
        bml_cuotas = ""

    row = {
        "StyleColor": str(base.get("StyleColor") or "").strip(),
        "ProductCode": str(base.get("ProductCode") or "").strip(),
        "Marketing Name": str(base.get("Marketing Name") or "").strip(),
        "Category": str(base.get("Category") or "").strip(),
        "Division": str(base.get("Division") or "").strip(),
        "Franchise": str(base.get("Franchise") or "").strip(),
        "Gender": str(base.get("Gender") or "").strip(),
        "Link PDP Competitor": str(meli.get("url") or "") if meli else "",
        "MeLi Buyable": (1 if (meli and meli.get("is_buyable")) else 0),
        "MeLi Status": str(meli.get("status") or "") if meli else "",
        "Competitor Full Price": format_money(comp_full),
        "Competitor Markdown": format_md(comp_md),
        "Competitor Final Price": format_money(comp_final),
        "Nike Full Price": format_money(nike_full),
        "Nike Markdown": f"{nike_md:.3f}",
        "Nike Final Price": format_money(nike_final),
        "Competitor vs Nike": comp_vs_nike,
        "BML Final Price": bml_final,
        "BML Full Price": bml_full,
        "Competitor Shipping": comp_ship,
        "Nike Shipping": nike_ship,
        "Nike Price + Shipping": nike_plus_ship,
        "Competitor Price + Shipping": comp_plus_ship,
        "BML with Shipping": bml_ship,
        "Cuotas Competitor": int(comp_cuotas or 0),
        "Cuotas Nike": int(nike_cuotas or 0),
        "BML Cuotas": bml_cuotas,
        "Competitor": COMPETITOR_NAME,
        "Season": output_season,
        "Fecha Corrida": run_ts,
        "Last Update Competitor": run_ts.split("_")[0],
    }
    return row

# ============================================================
# EXCEL WRITER
# ============================================================

FILL_BEAT = PatternFill("solid", fgColor="C6EFCE")
FILL_MEET = PatternFill("solid", fgColor="FFEB9C")
FILL_LOSE = PatternFill("solid", fgColor="FFC7CE")

def _bml_fill(v: str):
    v = (v or "").strip().lower()
    if v == "beat":
        return FILL_BEAT
    if v == "meet":
        return FILL_MEET
    if v == "lose":
        return FILL_LOSE
    return None

def write_xlsx(df: pd.DataFrame, out_path: str):
    """Escribe Excel con formato y colores BML."""
    log_info(f"Generando Excel: {out_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "MeLi vs Nike"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="F37021")
    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(df.shape[0]):
        for c, col in enumerate(df.columns, start=1):
            v = df.iloc[r, c-1]
            ws.cell(row=r+2, column=c, value=v)

    for c, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(x)) for x in df[col].astype(str).tolist()[:200]])
        ws.column_dimensions[get_column_letter(c)].width = min(55, max(12, max_len + 2))

    bml_cols = ["BML Final Price", "BML Full Price", "BML Cuotas", "BML with Shipping"]
    for bcol in bml_cols:
        if bcol in df.columns:
            col_idx = df.columns.get_loc(bcol) + 1
            for r in range(2, df.shape[0] + 2):
                v = ws.cell(row=r, column=col_idx).value
                fill = _bml_fill(v)
                if fill:
                    ws.cell(row=r, column=col_idx).fill = fill
                    ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal="center")

    wb.save(out_path)
    log_success(f"Excel guardado: {out_path}")

# ============================================================
# WORKERS
# ============================================================

def split_round_robin(items: List[Dict[str, Any]], n: int) -> List[List[Dict[str, Any]]]:
    """Distribuye items en n buckets (round-robin)."""
    if n <= 1:
        return [items]
    buckets = [[] for _ in range(n)]
    for i, it in enumerate(items):
        buckets[i % n].append(it)
    return [b for b in buckets if b]

def build_browser(playwright, headless: bool, country: str = "ar", session_id: Optional[str] = None, username_variant: Optional[str] = None, forced_port: Optional[int] = None):
    """Crea browser + context + page con proxy Decodo."""
    proxy = decodo_build_proxy(session_id=session_id, country=country, username_variant=username_variant, forced_port=forced_port)
    
    if not proxy:
        raise RuntimeError("❌ No se pudo configurar proxy Decodo")

    log_debug(f"Iniciando browser con proxy {proxy['server']} (usuario: {proxy['username']}, sesión: {session_id})")
    
    browser = playwright.chromium.launch(
        headless=headless,
        proxy=proxy,
        args=["--ignore-certificate-errors", "--disable-web-security"]
    )
    
    context = browser.new_context(
        locale="es-AR",
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        ignore_https_errors=True,
        viewport={"width": 1920, "height": 1080}
    )
    
    def _route_block_heavy(route, request):
        try:
            rtype = request.resource_type
        except Exception:
            rtype = None
        if rtype in ("image", "media", "font", "stylesheet"):
            return route.abort()
        return route.continue_()

    context.route("**/*", _route_block_heavy)
    page = context.new_page()
    page.set_default_timeout(DEFAULT_NAV_TIMEOUT_MS)
    
    return browser, context, page, proxy

def rotate_ip(current_browser, current_context, current_page, playwright, headless: bool, country: str = "ar", username_variant: Optional[str] = None, forced_port: Optional[int] = None):
    """Rota IP cerrando contexto y creando nueva sesión."""
    log_proxy("Rotando IP por detección de bloqueo...")
    
    try:
        current_page.close()
    except Exception:
        pass
    try:
        current_context.close()
    except Exception:
        pass
    try:
        current_browser.close()
    except Exception:
        pass

    time.sleep(ROTATION_BACKOFF_MS / 1000.0)

    new_session = uuid.uuid4().hex[:12]
    log_proxy(f"Nueva sesión: {new_session}")
    
    browser, context, page, proxy = build_browser(
        playwright,
        headless=headless,
        country=country,
        session_id=new_session,
        username_variant=username_variant,
        forced_port=forced_port
    )
    
    return browser, context, page, proxy

def worker_process_chunk(
    chunk: List[Dict[str, Any]],
    headless: bool,
    refresh_cached: bool,
    cache_subset: Dict[str, Any],
    worker_tag: str,
    username_variant: str,
    cache_obj: Cache,
    total_workers: int
) -> Tuple[List[Dict[str, Any]], Dict[str, Any], List[Dict[str, Any]]]:
    """Procesa un chunk de productos en un proceso aislado."""
    import time  # Garantiza disponibilidad en multiprocessing
    results_rows = []
    failed = []
    cache_updates = {}
    
    log_info(f"Worker {worker_tag} iniciando con {len(chunk)} productos")

    # --- Puerto fijo por worker + rotación round-robin (evita colisiones) ---
    try:
        worker_index = int(re.sub(r"\D+", "", worker_tag))  # "W3" -> 3
    except Exception:
        worker_index = 1

    # base_idx 0..(total_workers-1)
    base_idx = max(0, worker_index - 1)
    port_cursor = base_idx  # se moverá en saltos de total_workers
    def _current_forced_port() -> int:
        return DECODO_PORTS[int(port_cursor) % len(DECODO_PORTS)]


    with sync_playwright() as p:
        base_session = f"meli_{worker_tag}_{uuid.uuid4().hex[:8]}"
        browser, context, page, proxy = build_browser(
            p,
            headless=headless,
            country="ar",
            session_id=base_session,
            username_variant=username_variant,
            forced_port=_current_forced_port()
        )

        def rotate_current_ip():
            nonlocal browser, context, page, proxy
            nonlocal port_cursor
            port_cursor += int(total_workers)
            browser, context, page, proxy = rotate_ip(
                browser,
                context,
                page,
                p,
                headless=headless,
                country="ar",
                username_variant=username_variant,
                forced_port=_current_forced_port()
            )

        try:
            for idx, base in enumerate(chunk, start=1):
                mla = base["MLA"]
                nike_full = float(base.get("Nike Full Price Num") or 0.0)  # <--- VARIABLE CORRECTA

                # Desincronización suave entre workers (evita picos de conexiones simultáneas)
                try:
                    time.sleep(random.uniform(0.30, 1.00))
                except Exception:
                    pass

                # Verificar cache
                meli = None
                if (not refresh_cached) and mla in cache_subset:
                    cached = cache_subset[mla]
                    st_cached = (cached.get("status") or "")
                    if st_cached not in {"login_required", "account_verification"}:
                        meli = cached
                        log_debug(f"  [{worker_tag}] Cache hit para {mla}")

                if meli is None:
                    log_scraping(f"[{worker_tag} {idx}/{len(chunk)}] PDP {mla}")
                    
                    # Usar la función con reintentos - PASAMOS nike_full
                    meli = scrape_meli_pdp_with_retry(page, mla, nike_full)
                    
                    # Actualizar cache
                    cache_updates[mla] = meli

                    # Verificar si requiere rotación / rebuild
                    st = (meli.get("status") or "")
                    if (st in {"login_required", "account_verification"} and ROTATE_IP_ON_WALL):
                        log_warning(f"[{worker_tag}] Bloqueo detectado, rotando IP")
                        rotate_current_ip()
                    elif (meli.get("needs_rebuild") is True and REBUILD_ON_TUNNEL_ERRORS):
                        log_warning(f"[{worker_tag}] Túnel/proxy roto detectado, reconstruyendo browser y rotando puerto")
                        rotate_current_ip()
                    elif (st in {"proxy_block"} and ROTATE_IP_ON_WALL):
                        log_warning(f"[{worker_tag}] Proxy/IP bloqueado (403/429), rotando IP/puerto")
                        rotate_current_ip()

                # Solo activos van a output
                if meli and (meli.get("is_active") is True):
                    results_rows.append({"base": base, "meli": meli})
                    log_success(f"[{worker_tag}] {mla} procesado OK - ${meli.get('final_price_num', 0):,.2f}")
                elif meli:
                    # Registrar fallos no-activos para estadísticas
                    if meli.get("status") in ["not_found", "forbidden", "rate_limited", "server_error"]:
                        failed.append({
                            "MLA": mla, 
                            "error": meli.get("status", "unknown"), 
                            "detail": meli.get("error_detail", "")
                        })
                        log_error(f"[{worker_tag}] Falló {mla}: {meli.get('status', 'unknown')}")


        finally:
            try:
                page.close()
            except Exception:
                pass
            try:
                context.close()
            except Exception:
                pass
            try:
                browser.close()
            except Exception:
                pass

    log_info(f"Worker {worker_tag} completado: {len(results_rows)} OK, {len(failed)} fallos")
    return results_rows, cache_updates, failed

# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="MeLi vs Nike Scraper con Reintentos y Cache de 4 días")
    parser.add_argument("--statusbook", default=DEFAULT_STATUSBOOK_PATH)
    parser.add_argument("--season", default=DEFAULT_SEASON)
    parser.add_argument("--limit", type=int, default=0, help="Limitar productos (0 = todos)")
    parser.add_argument("--headless", action="store_true", default=DEFAULT_HEADLESS)
    parser.add_argument("--cache", default=CACHE_DEFAULT_PATH)
    parser.add_argument("--refresh-cache", action="store_true", default=REFRESH_CACHED_DEFAULT)
    parser.add_argument("--no-refresh-cache", action="store_true", default=False)
    parser.add_argument("--out-prefix", default="meli_vs_nike")
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS)
    parser.add_argument("--force-workers", action="store_true")
    parser.add_argument("--debug", action="store_true", default=True)
    parser.add_argument("--username-variant", type=str, default=None, help="Forzar variante de username")
    parser.add_argument("--ttl-days", type=int, default=CACHE_TTL_DAYS, help="Días de TTL para cache")
    args = parser.parse_args()

    global DEBUG_MODE
    DEBUG_MODE = args.debug

    season = args.season
    run_ts = now_ts()
    refresh_cached = args.refresh_cache and (not args.no_refresh_cache)

    print("\n" + "=" * 90)
    print("🚀 MeLi vs Nike — Southbay GOD Edition (Reintentos + Cache 4 días + 5 workers)")
    print("=" * 90)
    print(f"📅 Season: {season}")
    print(f"📚 StatusBook: {args.statusbook}")
    print(f"🧪 Headless: {args.headless} | Refresh Cache: {refresh_cached}")
    print(f"🗃️ Cache: {args.cache} (TTL: {args.ttl_days} días)")
    print(f"🔄 Reintentos: {MAX_SCRAPE_RETRIES} (backoff {RETRY_BACKOFF_FACTOR}^n)")
    print(f"👷 Workers: {args.workers}")
    print(f"🔌 Proxy Decodo Mobile: {DECODO_HOST} | Endpoints: {len(DECODO_PORTS)} | Ports: {DECODO_PORTS[0]}–{DECODO_PORTS[-1]}")
    print(f"👤 Usuario base: {DECODO_USER}")
    print("=" * 90 + "\n")

    # Test híbrido del proxy
    log_info("Iniciando test híbrido de proxy (Playwright + urllib3)...")
    
    if args.username_variant:
        log_info(f"Usando variante forzada: {args.username_variant}")
        success, data, method, username = test_proxy_playwright("ar", args.username_variant)
        if not success:
            success, data, method, username = test_proxy_urllib3("ar", args.username_variant)
        if not success:
            success, data, method, username = test_proxy_requests("ar", args.username_variant)
        working_username = args.username_variant if success else None
    else:
        success, data, method, working_username = decodo_test_connection_hybrid("ar")
    
    if success and data:
        log_success(f"✅ Proxy funcionando correctamente!")
        log_info(f"   Método: {method}")
        log_info(f"   Username: {working_username}")
        log_info(f"   IP: {data.get('proxy', {}).get('ip', 'N/A')}")
        log_info(f"   País: {data.get('country', {}).get('name', 'N/A')}")
    else:
        log_error("❌ Proxy NO responde después de probar todas las variantes.")
        respuesta = input("\n¿Continuar de todas formas (sin proxy)? (s/n): ")
        if respuesta.lower() != 's':
            return
        working_username = None

    # Leer universo
    try:
        universe = read_statusbook_universe(args.statusbook, season)
    except Exception as e:
        log_error(f"Error leyendo StatusBook: {e}")
        return

    if args.limit and args.limit > 0:
        universe = universe[:args.limit]
        log_info(f"Limitado a {args.limit} productos")

    # Inicializar cache con TTL configurable
    cache = Cache(args.cache, ttl_days=args.ttl_days)

    # Procesar
    all_results = []
    failed_items = []
    cache_updates_total = {}

    workers = max(1, int(args.workers))
    if workers > MAX_SAFE_WORKERS and not args.force_workers:
        log_warning(f"Workers={workers} > {MAX_SAFE_WORKERS}. Limitando a {MAX_SAFE_WORKERS}.")
        workers = MAX_SAFE_WORKERS

    if workers == 1:
        # Modo single-worker
        log_info("Modo single-worker")
        
        with sync_playwright() as p:
            base_session = f"meli_{uuid.uuid4().hex[:8]}"
            browser, context, page, proxy = build_browser(
                p, headless=args.headless, country="ar", session_id=base_session, username_variant=working_username
            )

            def rotate_current_ip():
                nonlocal browser, context, page, proxy
                browser, context, page, proxy = rotate_ip(
                    browser, context, page, p, headless=args.headless, country="ar", username_variant=working_username
                )

            try:
                for idx, base in enumerate(universe, start=1):
                    mla = base["MLA"]
                    nike_full = float(base.get("Nike Full Price Num") or 0.0)

                    # Usar cache
                    cached = cache.get(mla)
                    meli = None
                    if (not refresh_cached) and cached:
                        st_cached = (cached.get("status") or "")
                        if st_cached not in {"login_required", "account_verification"}:
                            meli = cached
                            log_debug(f"Cache hit para {mla}")

                    if meli is None:
                        log_scraping(f"[{idx}/{len(universe)}] PDP {mla}")
                        meli = scrape_meli_pdp_with_retry(page, mla, nike_full)
                        cache.set(mla, meli)
                        cache_updates_total[mla] = meli

                        if meli.get("status") in {"login_required", "account_verification"} and ROTATE_IP_ON_WALL:
                            log_warning(f"Bloqueo detectado, rotando IP")
                            rotate_current_ip()

                    if meli and (meli.get("is_active") is True):
                        all_results.append({"base": base, "meli": meli})
                        log_success(f"✓ {mla} - ${meli.get('final_price_num', 0):,.2f}")
                    elif meli:
                        failed_items.append({
                            "MLA": mla, 
                            "error": meli.get("status", "unknown"),
                            "detail": meli.get("error_detail", "")
                        })

            finally:
                try:
                    page.close()
                except Exception:
                    pass
                try:
                    context.close()
                except Exception:
                    pass
                try:
                    browser.close()
                except Exception:
                    pass

        cache.save()

    else:
        # Modo multi-worker
        log_info(f"Modo multi-worker con {workers} workers")
        chunks = split_round_robin(universe, workers)
        futures = []

        with ProcessPoolExecutor(max_workers=workers) as executor:
            for wi, chunk in enumerate(chunks, start=1):
                cache_subset = {it["MLA"]: cache.cache[it["MLA"]] for it in chunk if it["MLA"] in cache.cache}
                tag = f"W{wi}"
                log_info(f"Lanzando worker {tag} con {len(chunk)} productos")
                
                futures.append(
                    executor.submit(
                        worker_process_chunk,
                        chunk,
                        args.headless,
                        refresh_cached,
                        cache_subset,
                        tag,
                        working_username,
                        cache,
                        workers
                    )
                )

            for fut in as_completed(futures):
                rows, cache_updates, fails = fut.result()
                all_results.extend(rows)
                failed_items.extend(fails)
                cache_updates_total.update(cache_updates)

        # Actualizar cache con resultados de workers
        for mla, value in cache_updates_total.items():
            cache.set(mla, value)
        cache.save()

    # Construir output
    log_info(f"Construyendo {len(all_results)} filas de output...")
    
    output_rows = []
    for res in all_results:
        base = res["base"]
        meli = res["meli"]
        row = build_output_row(base, meli, season, run_ts)
        output_rows.append(row)

    df = pd.DataFrame(output_rows)
    
    for c in OUTPUT_COLS:
        if c not in df.columns:
            df[c] = ""
    df = df[OUTPUT_COLS]

    out_prefix = args.out_prefix
    csv_path = f"{out_prefix}_{run_ts}.csv"
    xlsx_path = f"{out_prefix}_{run_ts}.xlsx"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    write_xlsx(df, xlsx_path)

    print("\n" + "=" * 90)
    print("✅ PIPELINE COMPLETADO")
    print("=" * 90)
    print(f"📄 CSV: {csv_path}")
    print(f"📄 XLSX: {xlsx_path}")
    print(f"📊 Productos en output: {len(all_results)}")
    print(f"⚠️ Fallos: {len(failed_items)}")
    print(f"📦 Cache actualizado: {cache.size()} entradas (TTL: {args.ttl_days} días)")
    
    if failed_items:
        print("\nPrimeros 5 fallos:")
        for f in failed_items[:5]:
            error_detail = f.get('detail', '')
            if error_detail:
                print(f"   - {f['MLA']}: {f['error']} - {error_detail[:100]}")
            else:
                print(f"   - {f['MLA']}: {f['error']}")
    
    print("=" * 90 + "\n")

if __name__ == "__main__":
    main()